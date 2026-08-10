"""Batch IP reachability and reverse-DNS panel."""

from __future__ import annotations

import csv
import ipaddress
import re
import socket
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from PySide6.QtCore import QObject, Qt, Signal
from PySide6.QtGui import QColor, QFont
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QProgressBar,
    QPushButton,
    QSizePolicy,
    QSpinBox,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from quickip.ui_qt.tool_panels.components import allow_horizontal_shrink
from quickip.ui_qt.tool_panels.layout import configure_tool_root
from quickip.ui_qt.palette import semantic_color
from quickip.ui_qt.widgets.copyable_views import (
    CopyableTable,
    tree_selection_stylesheet,
)


_BATCH_MAX_WORKERS = 100
_BATCH_DEFAULT_WORKERS = 50


def _safe_workers(n_ips: int) -> int:
    """Return a safe worker count for *n_ips* addresses.

    Rules:
    - Never more than _BATCH_MAX_WORKERS (subprocess limit)
    - Never more than the number of IPs (no idle threads)
    - Scale down for small batches: no point in 50 threads for 10 IPs
    """
    return min(_BATCH_MAX_WORKERS, max(1, min(n_ips, _BATCH_DEFAULT_WORKERS)))


class IpBatchBridge(QObject):
    row_done   = Signal(int, str, str, str)   # row_idx, status, ms, hostname
    finished   = Signal(int, int)             # total, ok_count
    progress   = Signal(int, int, int)        # done, total, ok_count


def _ping_one(ip: str, timeout: int = 1, runner=None) -> tuple[bool, str]:
    """Ping *ip* once. Returns (reachable, rtt_ms_str)."""
    try:
        cmd = ["ping", "-n", "1", "-w", str(timeout * 1000), ip]
        r = runner.run(cmd, timeout=timeout + 2)
        if r.success:
            m = re.search(r"[Вв]ремя[<=](\d+)\s*мс|[Tt]ime[<=](\d+)\s*ms|[Tt]ime=(\d+)ms", r.stdout)
            ms = m.group(1) or m.group(2) or m.group(3) if m else "0"
            return True, ms
    except Exception:
        pass
    return False, ""


def _resolve_one(ip: str) -> str:
    """Reverse-DNS lookup for *ip*. Returns hostname or empty string."""
    try:
        return socket.gethostbyaddr(ip)[0]
    except Exception:
        return ""


class IpBatchPanel(QWidget):
    """Batch ping + reverse-DNS for a list of IPs loaded from CSV/Excel."""

    def __init__(self, dark: bool = True, i18n=None, runner=None) -> None:
        super().__init__()
        self._dark = dark
        self._i18n = i18n
        self._runner = runner
        self._rows: list[dict] = []      # original rows from file
        self._headers: list[str] = []    # column names
        self._ip_col: str = ""           # selected IP column
        self._stop_flag = False
        self._done_count = 0
        self._bridge = IpBatchBridge()
        self._bridge.row_done.connect(self._on_row_done)
        self._bridge.finished.connect(self._on_finished)
        self._bridge.progress.connect(self._on_progress)
        self._build()

    def _tr(self, key: str, **kwargs) -> str:
        s = self._i18n.get(key) if self._i18n else key
        return s.format(**kwargs) if kwargs else s

    # ── UI ────────────────────────────────────────────────────────

    @staticmethod
    def _step_label(n: int, text: str) -> QLabel:
        lbl = QLabel(f"  {n}. {text}")
        lbl.setStyleSheet(
            f'color: {semantic_color("TEXT_MUTED")}; font-size: 11px; font-weight: 600;'
        )
        return lbl

    @staticmethod
    def _divider() -> QFrame:
        d = QFrame()
        d.setFrameShape(QFrame.Shape.HLine)
        d.setObjectName("ToolDivider")
        return d

    def _build(self) -> None:
        _f = QFont()
        _f.setWeight(QFont.Weight.DemiBold)

        root = QVBoxLayout(self)
        configure_tool_root(root, spacing=12)

        self._title = QLabel(self._tr("tools_batch_title"))
        self._title.setObjectName("ToolPanelTitle")
        root.addWidget(self._title)

        source_card = QFrame()
        source_card.setObjectName("IpBatchSourceCard")
        source = QGridLayout(source_card)
        source.setContentsMargins(16, 14, 16, 16)
        source.setHorizontalSpacing(16)
        source.setVerticalSpacing(8)
        source.setColumnStretch(0, 0)
        source.setColumnStretch(1, 1)
        source.setColumnStretch(2, 1)
        source.setColumnStretch(3, 0)
        source.setColumnStretch(4, 0)

        self._lbl_step1 = self._card_label("tools_batch_source")
        self._lbl_step2 = self._card_label("tools_batch_col_label")
        self._lbl_step3 = self._card_label("tools_batch_timeout_label")
        self._lbl_workers = self._card_label("tools_batch_workers_label")
        source.addWidget(self._lbl_step1, 0, 0, 1, 2)
        source.addWidget(self._lbl_step2, 0, 2)
        source.addWidget(self._lbl_step3, 0, 3)
        source.addWidget(self._lbl_workers, 0, 4)

        self._btn_open = QPushButton(self._tr("tools_batch_open_btn"))
        self._btn_open.setProperty("role", "primary")
        self._btn_open.setObjectName("ToolBtn")
        self._btn_open.setMinimumSize(120, 40)
        self._btn_open.setMaximumWidth(220)
        self._btn_open.setFont(_f)
        self._btn_open.clicked.connect(self._on_open)

        self._lbl_file = QLabel(self._tr("tools_batch_no_file"))
        self._lbl_file.setObjectName("IpBatchFileLabel")
        self._lbl_file.setMinimumHeight(40)
        self._lbl_file.setMinimumWidth(140)
        self._lbl_file.setMaximumWidth(190)
        self._lbl_file.setSizePolicy(
            QSizePolicy.Policy.Preferred,
            QSizePolicy.Policy.Fixed,
        )

        self._col_combo = QComboBox()
        self._col_combo.setObjectName("ToolCombo")
        self._col_combo.setMinimumSize(190, 40)
        self._col_combo.setEnabled(False)
        self._col_combo.currentTextChanged.connect(self._on_col_changed)

        self._ed_timeout = QSpinBox()
        self._ed_timeout.setObjectName("ToolSpin")
        self._ed_timeout.setRange(1, 10)
        self._ed_timeout.setValue(1)
        self._ed_timeout.setSuffix(self._tr("tools_batch_seconds_suffix"))
        self._ed_timeout.setMinimumSize(90, 40)
        self._ed_timeout.setMaximumWidth(170)
        self._ed_workers = QSpinBox()
        self._ed_workers.setObjectName("ToolSpin")
        self._ed_workers.setRange(1, _BATCH_MAX_WORKERS)
        self._ed_workers.setValue(_BATCH_DEFAULT_WORKERS)
        self._ed_workers.setMinimumSize(80, 40)
        self._ed_workers.setMaximumWidth(150)

        source.addWidget(self._btn_open, 1, 0)
        source.addWidget(self._lbl_file, 1, 1)
        source.addWidget(self._col_combo, 1, 2)
        source.addWidget(self._ed_timeout, 1, 3)
        source.addWidget(self._ed_workers, 1, 4)
        for widget in (
            self._lbl_step1, self._lbl_step2, self._lbl_step3,
            self._lbl_workers,
            self._col_combo,
        ):
            allow_horizontal_shrink(widget)
        root.addWidget(source_card)

        actions_bar = QWidget(self)
        actions = QHBoxLayout(actions_bar)
        actions.setContentsMargins(0, 0, 0, 0)
        actions.setSpacing(8)
        self.btn_run = QPushButton(self._tr("tools_batch_run_btn"))
        self.btn_run.setProperty("role", "primary")
        self.btn_run.setObjectName("ToolBtn")
        self.btn_run.setMinimumSize(120, 40)
        self.btn_run.setFont(_f)
        self.btn_run.setEnabled(False)
        self.btn_run.clicked.connect(self._on_run)
        self.btn_stop = QPushButton(self._tr("tools_batch_stop_btn"))
        self.btn_stop.setProperty("role", "action")
        self.btn_stop.setObjectName("ToolBtn")
        self.btn_stop.setMinimumSize(110, 40)
        self.btn_stop.setFont(_f)
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self._on_stop)

        self._search = QLineEdit()
        self._search.setObjectName("ToolInput")
        self._search.setPlaceholderText(self._tr("tools_batch_search_placeholder"))
        self._search.setClearButtonEnabled(True)
        self._search.setMinimumHeight(40)
        self._search.textChanged.connect(self._apply_filter)

        self._result_filter = QComboBox()
        self._result_filter.setObjectName("ToolCombo")
        self._result_filter.setMinimumSize(160, 40)
        for key in ("all", "reachable", "unreachable", "invalid", "pending"):
            self._result_filter.addItem(self._tr(f"tools_batch_filter_{key}"), key)
        self._result_filter.currentIndexChanged.connect(self._apply_filter)

        self._status = QLabel("")
        self._status.setObjectName("ToolStatus")
        actions.addWidget(self.btn_run)
        actions.addWidget(self.btn_stop)
        actions.addWidget(self._search, 1)
        actions.addWidget(self._result_filter)
        # The search field absorbs horizontal compression while buttons and
        # the result filter retain readable widths.
        allow_horizontal_shrink(self._search)
        root.addWidget(actions_bar)

        summary = QHBoxLayout()
        summary.setSpacing(8)
        self._total_pill = self._summary_pill()
        self._ok_pill = self._summary_pill()
        self._failed_pill = self._summary_pill()
        self._invalid_pill = self._summary_pill()
        for pill in (self._total_pill, self._ok_pill, self._failed_pill, self._invalid_pill):
            summary.addWidget(pill)
        summary.addStretch(1)
        root.addLayout(summary)

        progress_row = QFrame()
        progress_row.setObjectName("IpBatchProgressRow")
        progress_layout = QHBoxLayout(progress_row)
        progress_layout.setContentsMargins(12, 8, 12, 8)
        progress_layout.setSpacing(12)
        self._progress = QProgressBar()
        self._progress.setObjectName("IpBatchProgress")
        self._progress.setRange(0, 100)
        self._progress.setValue(0)
        self._progress.setTextVisible(False)
        self._progress.setFixedHeight(7)
        self._progress.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Fixed,
        )
        self._status.setText(self._tr("tools_batch_ready"))
        self._status.setMinimumWidth(210)
        self._status.setAlignment(
            Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
        )
        progress_layout.addWidget(self._progress, 1)
        progress_layout.addWidget(self._status)
        root.addWidget(progress_row)

        self._table = CopyableTable(0, 0, i18n=self._i18n)
        self._table.setObjectName("IpBatchTable")
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self._table.verticalHeader().setDefaultSectionSize(36)
        self._table.setSortingEnabled(True)
        self._show_empty_table_columns()
        root.addWidget(self._table, 1)

        footer = QHBoxLayout()
        self._btn_export = QPushButton(self._tr("tools_batch_save_btn"))
        self._btn_export.setProperty("role", "action")
        self._btn_export.setObjectName("ToolBtn")
        self._btn_export.setMinimumSize(180, 40)
        self._btn_export.setFont(_f)
        self._btn_export.setEnabled(False)
        self._btn_export.clicked.connect(self._on_export)
        footer.addStretch(1)
        footer.addWidget(self._btn_export)
        root.addLayout(footer)
        self._update_summary()
    def _card_label(self, key: str) -> QLabel:
        label = QLabel(self._tr(key))
        label.setObjectName("IpBatchFieldLabel")
        label.setWordWrap(True)
        label.setMinimumHeight(32)
        return label

    @staticmethod
    def _summary_pill() -> QLabel:
        label = QLabel()
        label.setObjectName("IpBatchSummaryPill")
        label.setMinimumHeight(26)
        label.setMinimumWidth(90)
        label.setMaximumWidth(160)
        label.setSizePolicy(
            QSizePolicy.Policy.Preferred,
            QSizePolicy.Policy.Fixed,
        )
        return label

    # ── File loading ─────────────────────────────────────────────

    def _on_open(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, self._tr("tools_batch_dlg_open"),
            "", self._tr("tools_batch_dlg_filter")
        )
        if not path:
            return
        try:
            self._rows, self._headers = self._load_file(path)
        except Exception as exc:
            self._status.setText(self._tr("tools_batch_err_read", err=exc))
            self._status.setStyleSheet(
                f'color: {semantic_color("STATUS_ERROR")}; font-size: 12px;'
            )
            return

        self._lbl_file.setText(self._tr("tools_batch_loaded", name=Path(path).name, n=len(self._rows)))
        self._col_combo.setEnabled(True)
        self._col_combo.blockSignals(True)
        self._col_combo.clear()
        self._col_combo.addItems(self._headers)
        # auto-select first column that looks like IPs
        auto = next(
            (h for h in self._headers if re.search(r"ip|addr|хост|host", h, re.IGNORECASE)),
            self._headers[0] if self._headers else "",
        )
        self._col_combo.setCurrentText(auto)
        self._col_combo.blockSignals(False)
        self._ip_col = self._col_combo.currentText()

        self._build_table(self._rows, self._headers, self._ip_col)
        self.btn_run.setEnabled(bool(self._rows))
        self._btn_export.setEnabled(False)
        self._progress.setValue(0)
        # Auto-tune worker count based on batch size
        self._ed_workers.setValue(_safe_workers(len(self._rows)))
        self._status.setText(self._tr("tools_batch_status_loaded", n=len(self._rows)))
        self._status.setStyleSheet(
            f'color: {semantic_color("TEXT_MUTED")}; font-size: 12px;'
        )
        self._update_summary()

    @staticmethod
    def _load_file(path: str) -> tuple[list[dict], list[str]]:
        p = path.lower()
        if p.endswith(".csv"):
            # Try utf-8-sig first, fall back to cp1251 for Windows-exported files
            for enc in ("utf-8-sig", "cp1251", "utf-8"):
                try:
                    with open(path, newline="", encoding=enc) as f:
                        sample = f.read(4096)
                    # Auto-detect delimiter: count candidates in header line
                    first_line = sample.splitlines()[0] if sample else ""
                    dialect_delim = max(",", ";", "\t", "|",
                                        key=lambda d: first_line.count(d))
                    with open(path, newline="", encoding=enc) as f:
                        reader = csv.DictReader(f, delimiter=dialect_delim)
                        rows = list(reader)
                        headers = list(reader.fieldnames) if reader.fieldnames else (
                            list(rows[0].keys()) if rows else []
                        )
                    return rows, headers
                except UnicodeDecodeError:
                    continue
            raise ValueError("Не удалось определить кодировку CSV")
        # Excel (.xlsx / .xls)
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError("openpyxl не найден (обратитесь к разработчику)") from exc
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return [], []
        headers = [str(c) if c is not None else f"col{i}" for i, c in enumerate(all_rows[0])]
        rows = [
            {headers[i]: (str(cell) if cell is not None else "") for i, cell in enumerate(row)}
            for row in all_rows[1:]
        ]
        return rows, headers

    def _on_col_changed(self, col: str) -> None:
        self._ip_col = col
        if self._rows:
            self._build_table(self._rows, self._headers, col)
            self._btn_export.setEnabled(False)

    def _show_empty_table_columns(self) -> None:
        """Show a useful table structure before a source file is loaded."""
        columns = [
            "IP",
            self._tr("tools_batch_col_status"),
            self._tr("tools_batch_col_rtt"),
            self._tr("tools_batch_col_hostname"),
        ]
        self._table.setSortingEnabled(False)
        self._table.setRowCount(0)
        self._table.setColumnCount(len(columns))
        self._table.setHorizontalHeaderLabels(columns)
        header = self._table.horizontalHeader()
        header.setDefaultAlignment(
            Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter
        )
        header.setStretchLastSection(False)
        for column in range(len(columns)):
            header.setSectionResizeMode(column, QHeaderView.ResizeMode.Stretch)
        self._table.setSortingEnabled(True)

    def _build_table(self, rows: list[dict], headers: list[str], ip_col: str) -> None:
        """Rebuild results table: original columns + Status, RTT, Hostname."""
        extra = [
            self._tr("tools_batch_col_status"),
            self._tr("tools_batch_col_rtt"),
            self._tr("tools_batch_col_hostname"),
        ]
        all_cols = headers + extra
        self._table.setSortingEnabled(False)
        self._table.setRowCount(0)
        self._table.setColumnCount(len(all_cols))
        self._table.setHorizontalHeaderLabels(all_cols)
        for column in range(len(all_cols)):
            header_item = self._table.horizontalHeaderItem(column)
            if header_item is not None:
                header_item.setTextAlignment(
                    Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter
                )
        self._table.setRowCount(len(rows))
        for r_idx, row in enumerate(rows):
            for c_idx, col in enumerate(headers):
                item = QTableWidgetItem(str(row.get(col, "")))
                if col == ip_col:
                    item.setFont(QFont("Consolas", 9))
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter
                )
                self._table.setItem(r_idx, c_idx, item)
            for c_idx, _ in enumerate(extra):
                item = QTableWidgetItem("")
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter
                )
                self._table.setItem(r_idx, len(headers) + c_idx, item)
        self._table.resizeColumnsToContents()
        table_header = self._table.horizontalHeader()
        # RU: Заголовки и данные центрируются внутри своих столбцов.
        # EN: Center headers and values inside their respective columns.
        table_header.setDefaultAlignment(
            Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter
        )
        table_header.setStretchLastSection(False)
        table_header.setMinimumSectionSize(90)

        status_col = len(headers)
        rtt_col = status_col + 1
        hostname_col = status_col + 2

        if len(all_cols) <= 6:
            # RU: В компактной таблице все столбцы образуют ровную сетку и
            # вместе занимают всю доступную ширину.
            # EN: In compact tables every column forms an even grid and the
            # columns collectively fill all available horizontal space.
            for c_idx in range(len(all_cols)):
                table_header.setSectionResizeMode(c_idx, QHeaderView.ResizeMode.Stretch)
        else:
            # RU: При большом числе столбцов сохраняем читаемую ширину и
            # горизонтальную прокрутку вместо чрезмерного сжатия текста.
            # EN: With many columns, preserve readable widths and horizontal
            # scrolling instead of compressing cell text excessively.
            for c_idx, col in enumerate(headers):
                table_header.setSectionResizeMode(
                    c_idx, QHeaderView.ResizeMode.Interactive
                )
                minimum_width = 155 if col == ip_col else 120
                measured_width = self._table.columnWidth(c_idx)
                self._table.setColumnWidth(
                    c_idx, max(minimum_width, min(measured_width, 280))
                )
            table_header.setSectionResizeMode(
                status_col, QHeaderView.ResizeMode.Interactive
            )
            table_header.setSectionResizeMode(
                rtt_col, QHeaderView.ResizeMode.Interactive
            )
            table_header.setSectionResizeMode(
                hostname_col, QHeaderView.ResizeMode.Stretch
            )
            self._table.setColumnWidth(status_col, 145)
            self._table.setColumnWidth(rtt_col, 110)
        self._table.setSortingEnabled(True)
        self._apply_filter()
        self._update_summary()

    # ── Run / Stop ────────────────────────────────────────────────

    def _on_run(self) -> None:
        if not self._ip_col or not self._rows:
            return
        self._stop_flag = False
        self._done_count = 0
        self._btn_export.setEnabled(False)
        self.btn_run.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self._table.setSortingEnabled(False)
        self._progress.setValue(0)
        # Reset result columns
        n_orig = len(self._headers)
        for r_idx in range(len(self._rows)):
            for c in range(3):
                item = QTableWidgetItem("")
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter
                )
                self._table.setItem(r_idx, n_orig + c, item)
        timeout = self._ed_timeout.value()
        workers = self._ed_workers.value()
        self._status.setText(self._tr("tools_batch_starting"))
        threading.Thread(
            target=self._worker,
            args=(list(self._rows), self._ip_col, self._headers, timeout, workers),
            daemon=True,
        ).start()

    def _on_stop(self) -> None:
        self._stop_flag = True

    def _worker(
        self, rows: list[dict], ip_col: str, headers: list[str], timeout: int, max_workers: int
    ) -> None:
        total = len(rows)
        ok_count = 0
        lock = threading.Lock()

        def _check(idx: int, row: dict):
            if self._stop_flag:
                return idx, "Pending", "", ""
            ip = str(row.get(ip_col, "")).strip()
            try:
                parsed = ipaddress.ip_address(ip)
                if parsed.version != 4:
                    raise ValueError
            except ValueError:
                return idx, "Invalid", "", ""
            if self._runner is None:
                return idx, "Error", "", ""
            reachable, ms = _ping_one(ip, timeout, runner=self._runner)
            hostname = _resolve_one(ip) if reachable else ""
            status = "OK" if reachable else "Timeout"
            return idx, status, ms, hostname

        # Hard cap: never exceed _BATCH_MAX_WORKERS regardless of UI value,
        # and never create more threads than there are IPs to check.
        effective = min(max_workers, _BATCH_MAX_WORKERS, len(rows))
        with ThreadPoolExecutor(max_workers=effective) as pool:
            futures = {pool.submit(_check, i, row): i for i, row in enumerate(rows)}
            for fut in as_completed(futures):
                if self._stop_flag:
                    pool.shutdown(wait=False, cancel_futures=True)
                    break
                idx, status, ms, hostname = fut.result()
                with lock:
                    if status == "OK":
                        ok_count += 1
                    done = self._done_count + 1
                    self._done_count = done
                self._bridge.row_done.emit(idx, status, ms, hostname)
                self._bridge.progress.emit(done, total, ok_count)

        self._bridge.finished.emit(total, ok_count)

    # ── Signals ───────────────────────────────────────────────────

    def _on_row_done(self, row_idx: int, status: str, ms: str, hostname: str) -> None:
        n = len(self._headers)
        ok = status == "OK"
        color = semantic_color(
            "STATUS_SUCCESS"
            if ok
            else ("STATUS_WARNING" if status in {"Invalid", "Pending"} else "STATUS_ERROR")
        )
        status_text = self._localized_status(status)
        for c_off, text in enumerate([status_text, ms, hostname]):
            item = QTableWidgetItem(text)
            item.setTextAlignment(
                Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter
            )
            if c_off == 0:
                item.setData(Qt.ItemDataRole.UserRole, status)
                item.setForeground(QColor(color))
                item.setFont(QFont("Consolas", 9))
            self._table.setItem(row_idx, n + c_off, item)
        self._apply_filter()
        self._update_summary()

    def _on_progress(self, done: int, total: int, ok: int) -> None:
        pct = int(done / total * 100) if total else 0
        self._progress.setValue(pct)
        self._status.setText(self._tr("tools_batch_progress", done=done, total=total, ok=ok))
        self._status.setStyleSheet(
            f'color: {semantic_color("TEXT_MUTED")}; font-size: 12px;'
        )

    def _on_finished(self, total: int, ok: int) -> None:
        self.btn_run.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self._btn_export.setEnabled(True)
        self._table.setSortingEnabled(True)
        stopped = self._stop_flag
        key = "tools_batch_stopped" if stopped else "tools_batch_done"
        msg = self._tr(key, ok=ok, total=total)
        color = semantic_color(
            "STATUS_SUCCESS" if ok == total and not stopped else "STATUS_WARNING"
        )
        self._status.setText(msg)
        self._status.setStyleSheet(f"color: {color}; font-size: 12px;")
        self._apply_filter()
        self._update_summary()

    def _localized_status(self, status: str) -> str:
        key = {
            "OK": "reachable",
            "Timeout": "unreachable",
            "Invalid": "invalid",
            "Pending": "pending",
            "Error": "error",
        }.get(status, "pending")
        return self._tr(f"tools_batch_status_{key}")

    def _raw_status(self, row: int) -> str:
        item = self._table.item(row, len(self._headers))
        return str(item.data(Qt.ItemDataRole.UserRole) or "Pending") if item else "Pending"

    def _apply_filter(self) -> None:
        if not hasattr(self, "_table"):
            return
        query = self._search.text().strip().casefold() if hasattr(self, "_search") else ""
        selected = (
            str(self._result_filter.currentData() or "all")
            if hasattr(self, "_result_filter")
            else "all"
        )
        mapping = {
            "reachable": {"OK"},
            "unreachable": {"Timeout", "Error"},
            "invalid": {"Invalid"},
            "pending": {"Pending"},
        }
        accepted = mapping.get(selected)
        for row in range(self._table.rowCount()):
            values = [
                self._table.item(row, col).text()
                for col in range(self._table.columnCount())
                if self._table.item(row, col)
            ]
            matches_query = not query or any(query in value.casefold() for value in values)
            matches_status = accepted is None or self._raw_status(row) in accepted
            self._table.setRowHidden(row, not (matches_query and matches_status))

    def _update_summary(self) -> None:
        if not hasattr(self, "_total_pill"):
            return
        statuses = [self._raw_status(row) for row in range(self._table.rowCount())]
        ok = statuses.count("OK")
        failed = sum(status in {"Timeout", "Error"} for status in statuses)
        invalid = statuses.count("Invalid")
        self._total_pill.setText(self._tr("tools_batch_summary_total", count=len(statuses)))
        self._ok_pill.setText(self._tr("tools_batch_summary_ok", count=ok))
        self._failed_pill.setText(self._tr("tools_batch_summary_failed", count=failed))
        self._invalid_pill.setText(self._tr("tools_batch_summary_invalid", count=invalid))

    # ── Export ────────────────────────────────────────────────────

    def _on_export(self) -> None:
        path, chosen = QFileDialog.getSaveFileName(
            self, self._tr("tools_batch_dlg_save"),
            "ip_check_results",
            self._tr("tools_batch_dlg_save_filter"),
        )
        if not path:
            return
        try:
            if chosen.startswith("Excel") or path.lower().endswith(".xlsx"):
                if not path.lower().endswith(".xlsx"):
                    path += ".xlsx"
                self._export_xlsx(path)
            else:
                if not path.lower().endswith(".csv"):
                    path += ".csv"
                self._export_csv(path)
            self._status.setText(self._tr("tools_batch_saved", name=Path(path).name))
            self._status.setStyleSheet(
                f'color: {semantic_color("STATUS_SUCCESS")}; font-size: 12px;'
            )
        except Exception as exc:
            self._status.setText(self._tr("tools_batch_err_save", err=exc))
            self._status.setStyleSheet(
                f'color: {semantic_color("STATUS_ERROR")}; font-size: 12px;'
            )

    def _collect_results(self) -> tuple[list[str], list[list[str]]]:
        """Read current table contents into (headers, data_rows)."""
        cols = self._table.columnCount()
        rows = self._table.rowCount()
        headers = [self._table.horizontalHeaderItem(c).text() for c in range(cols)]
        data = [
            [
                (self._table.item(r, c).text() if self._table.item(r, c) else "")
                for c in range(cols)
            ]
            for r in range(rows)
        ]
        return headers, data

    def _export_csv(self, path: str) -> None:
        headers, data = self._collect_results()
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data)

    def _export_xlsx(self, path: str) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError as exc:
            raise ImportError("openpyxl не найден (обратитесь к разработчику)") from exc
        headers, data = self._collect_results()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IP Check"
        # Header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")
        for c_idx, col in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        # Data rows
        status_col = len(headers) - 2  # "Статус" column (1-based)
        for r_idx, row in enumerate(data, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if c_idx == status_col:
                    if val == "OK":
                        cell.font = Font(color="22C55E", bold=True)
                    elif val == "Timeout":
                        cell.font = Font(color="EF4444")
        # Column widths
        for c_idx, col in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = max(
                len(col) + 2,
                max((len(str(row[c_idx - 1])) for row in data), default=0) + 2,
                10,
            )
        wb.save(path)

    def retranslate(self) -> None:
        self._title.setText(self._tr("tools_batch_title"))
        self._lbl_step1.setText(self._tr("tools_batch_source"))
        self._lbl_step2.setText(self._tr("tools_batch_col_label"))
        self._lbl_step3.setText(self._tr("tools_batch_timeout_label"))
        self._btn_open.setText(self._tr("tools_batch_open_btn"))
        self._lbl_workers.setText(self._tr("tools_batch_workers_label"))
        self._ed_timeout.setSuffix(self._tr("tools_batch_seconds_suffix"))
        self.btn_run.setText(self._tr("tools_batch_run_btn"))
        self.btn_stop.setText(self._tr("tools_batch_stop_btn"))
        self._btn_export.setText(self._tr("tools_batch_save_btn"))
        self._search.setPlaceholderText(self._tr("tools_batch_search_placeholder"))
        for index, key in enumerate(("all", "reachable", "unreachable", "invalid", "pending")):
            self._result_filter.setItemText(index, self._tr(f"tools_batch_filter_{key}"))
        if not self._rows:
            self._lbl_file.setText(self._tr("tools_batch_no_file"))
            self._show_empty_table_columns()
        self._update_summary()

    def refresh_theme(self, dark: bool) -> None:
        self._dark = dark
        self._table.setStyleSheet(tree_selection_stylesheet(dark))
