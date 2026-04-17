"""Tools page — ping, traceroute, DNS lookup, port scan, ipconfig, flush DNS."""

from __future__ import annotations

import base64
import csv
import json
import re
import socket
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import TYPE_CHECKING

from PySide6.QtCore import Qt, QObject, Signal, QSize
from PySide6.QtGui import QColor, QFont, QIcon, QPainter, QPen, QTextCharFormat
from PySide6.QtWidgets import (
    QAbstractItemView, QApplication, QComboBox, QFileDialog, QFrame, QGridLayout,
    QHBoxLayout, QHeaderView, QLabel, QLineEdit, QMenu, QMessageBox, QProgressBar,
    QPushButton, QSizePolicy, QSpinBox, QStackedWidget, QStyleFactory,
    QTableWidget, QTableWidgetItem, QTreeWidget, QTreeWidgetItem, QTextEdit,
    QVBoxLayout, QWidget,
)

if TYPE_CHECKING:
    from quickip.app.bootstrap import ServiceContainer


_TREE_SS_LIGHT = """
QTreeWidget::item:hover {
    background: #E0E7FF;
    color: #1E293B;
}
QTreeWidget::item:selected,
QTreeWidget::item:selected:active,
QTreeWidget::item:selected:!active {
    background: #6366F1;
    color: #FFFFFF;
}
"""

_TREE_SS_DARK = """
QTreeWidget::item:hover {
    background: rgba(99,102,241,0.18);
    color: #E2E8F0;
}
QTreeWidget::item:selected,
QTreeWidget::item:selected:active,
QTreeWidget::item:selected:!active {
    background: #4F46E5;
    color: #FFFFFF;
}
"""


class _CopyableTree(QTreeWidget):
    """QTreeWidget with Ctrl+C copy and right-click context menu."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)

    def keyPressEvent(self, event) -> None:
        mod = event.modifiers()
        key = event.key()
        if key == Qt.Key.Key_C and mod == Qt.KeyboardModifier.ControlModifier:
            self._copy_selection()
        elif key == Qt.Key.Key_A and mod == Qt.KeyboardModifier.ControlModifier:
            self.selectAll()
        else:
            super().keyPressEvent(event)

    def contextMenuEvent(self, event) -> None:
        item = self.itemAt(event.pos())
        if not item:
            return
        menu = QMenu(self)
        col = self.currentColumn()
        act_cell = menu.addAction("Копировать ячейку")
        act_row  = menu.addAction("Копировать строку")
        act_sel  = menu.addAction("Копировать выделенное")
        act_all  = menu.addAction("Копировать всё")
        chosen = menu.exec(event.globalPos())
        if chosen == act_cell:
            QApplication.clipboard().setText(item.text(col if col >= 0 else 0))
        elif chosen == act_row:
            texts = [item.text(c) for c in range(self.columnCount())]
            QApplication.clipboard().setText("\t".join(texts))
        elif chosen == act_sel:
            self._copy_selection()
        elif chosen == act_all:
            self.selectAll()
            self._copy_selection()

    def _copy_selection(self) -> None:
        items = self.selectedItems()
        if not items:
            return
        # selectedItems() returns all selected cells — group by row
        seen: dict[int, QTreeWidgetItem] = {}
        for it in items:
            idx = self.indexOfTopLevelItem(it)
            if idx >= 0:
                seen[idx] = it
        ordered = [seen[k] for k in sorted(seen)]
        lines = []
        for it in ordered:
            lines.append("\t".join(it.text(c) for c in range(self.columnCount())))
        QApplication.clipboard().setText("\n".join(lines))


class _CopyableTable(QTableWidget):
    """QTableWidget with Ctrl+C copy and right-click context menu."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)

    def keyPressEvent(self, event) -> None:
        mod = event.modifiers()
        key = event.key()
        if key == Qt.Key.Key_C and mod == Qt.KeyboardModifier.ControlModifier:
            self._copy_selection()
        elif key == Qt.Key.Key_A and mod == Qt.KeyboardModifier.ControlModifier:
            self.selectAll()
        else:
            super().keyPressEvent(event)

    def contextMenuEvent(self, event) -> None:
        index = self.indexAt(event.pos())
        if not index.isValid():
            return
        menu = QMenu(self)
        act_cell = menu.addAction("Копировать ячейку")
        act_row  = menu.addAction("Копировать строку")
        act_sel  = menu.addAction("Копировать выделенное")
        act_all  = menu.addAction("Копировать всё")
        chosen = menu.exec(event.globalPos())
        if chosen == act_cell:
            item = self.item(index.row(), index.column())
            QApplication.clipboard().setText(item.text() if item else "")
        elif chosen == act_row:
            texts = []
            for c in range(self.columnCount()):
                it = self.item(index.row(), c)
                texts.append(it.text() if it else "")
            QApplication.clipboard().setText("\t".join(texts))
        elif chosen == act_sel:
            self._copy_selection()
        elif chosen == act_all:
            self.selectAll()
            self._copy_selection()

    def _copy_selection(self) -> None:
        ranges = self.selectedRanges()
        if not ranges:
            return
        rows = sorted({r for rng in ranges for r in range(rng.topRow(), rng.bottomRow() + 1)})
        lines = []
        for row in rows:
            texts = []
            for c in range(self.columnCount()):
                it = self.item(row, c)
                texts.append(it.text() if it else "")
            lines.append("\t".join(texts))
        QApplication.clipboard().setText("\n".join(lines))


class _Bridge(QObject):
    output        = Signal(str, bool)
    finished      = Signal(bool, str)
    chart_update  = Signal(list)   # ping chart data


class _ToolPanel(QWidget):
    def __init__(self, title: str, dark: bool = True, i18n=None, runner=None) -> None:
        super().__init__()
        self._dark = dark
        self._i18n = i18n
        self._runner = runner
        self._bridge = _Bridge()
        self._running = False
        self._proc = None

        root = QVBoxLayout(self)
        root.setContentsMargins(16, 14, 16, 14)
        root.setSpacing(10)

        self._hdr = QLabel(title)
        self._hdr.setObjectName("ToolPanelTitle")
        root.addWidget(self._hdr)

        self._form = QHBoxLayout()
        self._form.setSpacing(8)
        self._form.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        root.addLayout(self._form)

        _f = QFont("Segoe UI", 10)
        _f.setWeight(QFont.Weight.DemiBold)
        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        btn_row.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        self.btn_run = QPushButton("\u25b6  Запуск")
        self.btn_run.setProperty("role", "primary")
        self.btn_run.setObjectName("ToolBtn")
        self.btn_run.setFixedSize(90, 28)
        self.btn_run.setFont(_f)
        self.btn_stop = QPushButton("\u25a0  Стоп")
        self.btn_stop.setProperty("role", "action")
        self.btn_stop.setObjectName("ToolBtn")
        self.btn_stop.setFixedSize(80, 28)
        self.btn_stop.setFont(_f)
        self.btn_stop.setEnabled(False)
        self.btn_clear = QPushButton("Очистить")
        self.btn_clear.setProperty("role", "action")
        self.btn_clear.setObjectName("ToolBtn")
        self.btn_clear.setFixedSize(90, 28)
        self.btn_clear.setFont(_f)
        btn_row.addWidget(self.btn_run)
        btn_row.addWidget(self.btn_stop)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_clear)
        root.addLayout(btn_row)

        self._output = QTextEdit()
        self._output.setObjectName("ToolOutput")
        self._output.setReadOnly(True)
        self._output.setFont(QFont("Consolas", 10))
        root.addWidget(self._output, 1)

        self._status = QLabel("")
        self._status.setObjectName("ToolStatus")
        root.addWidget(self._status)

        self._bridge.output.connect(self._on_output)
        self._bridge.finished.connect(self._on_finished)
        self.btn_run.clicked.connect(self._on_run)
        self.btn_stop.clicked.connect(self._on_stop)
        self.btn_clear.clicked.connect(self._output.clear)

    def _tr(self, key: str) -> str:
        return self._i18n.get(key) if self._i18n else key

    def _on_run(self) -> None:
        pass

    def _on_stop(self) -> None:
        self._running = False
        if self._proc:
            try:
                self._proc.kill()
            except Exception:
                pass
        self.btn_run.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self._status.setText("Остановлено")

    def _set_running(self, running: bool) -> None:
        self._running = running
        self.btn_run.setEnabled(not running)
        self.btn_stop.setEnabled(running)

    def _on_output(self, text: str, is_error: bool) -> None:
        if is_error:
            self._output.setTextColor(QColor("#EF4444"))
            self._output.append(text)
            self._output.setCurrentCharFormat(QTextCharFormat())
        else:
            self._output.setCurrentCharFormat(QTextCharFormat())
            self._output.append(text)
        self._output.verticalScrollBar().setValue(
            self._output.verticalScrollBar().maximum()
        )

    def _on_finished(self, success: bool, summary: str) -> None:
        self._set_running(False)
        color = "#22C55E" if success else "#EF4444"
        self._status.setText(summary)
        self._status.setStyleSheet(f"color: {color}; font-size: 12px;")

    def refresh_theme(self, dark: bool) -> None:
        self._dark = dark


class _PingChart(QWidget):
    def __init__(self, dark: bool = True) -> None:
        super().__init__()
        self._data: list = []
        self._dark = dark
        self.setMinimumHeight(40)

    def set_data(self, data: list) -> None:
        self._data = data
        self.update()

    def paintEvent(self, event) -> None:
        if not self._data:
            return
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w, h = self.width(), self.height()
        margin = 4
        valid = [d for d in self._data if d > 0]
        max_val = max(valid) if valid else 1
        bar_w = max(2, (w - margin * 2) // max(len(self._data), 1) - 1)
        for i, val in enumerate(self._data):
            x = margin + i * (bar_w + 1)
            if val == 0:
                color = QColor("#EF4444")
                bar_h = h - margin * 2
            else:
                ratio = val / max_val
                color = QColor("#22C55E") if ratio < 0.33 else (QColor("#F59E0B") if ratio < 0.66 else QColor("#EF4444"))
                bar_h = max(2, int((h - margin * 2) * ratio))
            p.setPen(Qt.PenStyle.NoPen)
            p.setBrush(color)
            p.drawRoundedRect(x, h - margin - bar_h, bar_w, bar_h, 2, 2)
        p.end()

    def refresh_theme(self, dark: bool) -> None:
        self._dark = dark
        self.update()


class _PingPanel(_ToolPanel):
    def __init__(self, dark: bool = True, runner=None) -> None:
        super().__init__("Ping", dark, runner=runner)
        self._host = QLineEdit()
        self._host.setObjectName("ToolInput")
        self._host.setPlaceholderText("Host or IP (e.g. 8.8.8.8)")
        self._host.setFixedHeight(28)
        self._host.returnPressed.connect(self._on_run)
        self._form.addWidget(self._host, 1)

        lbl = QLabel("Count:")
        lbl.setObjectName("FieldLabel")
        self._count = QSpinBox()
        self._count.setObjectName("ToolSpinBox")
        self._count.setRange(1, 100)
        self._count.setValue(3)
        self._count.setFixedWidth(50)
        self._count.setFixedHeight(28)
        self._count.setButtonSymbols(QSpinBox.ButtonSymbols.NoButtons)
        self._form.addWidget(lbl)
        self._form.addWidget(self._count)


    def _on_run(self) -> None:
        host = self._host.text().strip()
        if not host:
            self._status.setText("Введите хост")
            return
        self._output.clear()
        self._set_running(True)
        self._status.setText(f"Pinging {host}...")
        threading.Thread(target=self._worker, args=(host, self._count.value()), daemon=True).start()

    def _worker(self, host: str, count: int) -> None:
        try:
            cmd = ["ping", "-n", str(count), host]
            self._proc = self._runner.popen(cmd, encoding="cp866", errors="replace")
            times: list = []
            success_count = 0
            for line in self._proc.stdout or []:  # type: ignore[union-attr]
                if not self._running:
                    break
                line = line.rstrip()
                self._bridge.output.emit(line, False)
                m = re.search(r"[=<](\d+)\s*мс|time[=<](\d+)\s*ms", line, re.IGNORECASE)
                if m:
                    ms = float(m.group(1) or m.group(2))
                    times.append(ms)
                    success_count += 1
                    self._bridge.chart_update.emit(list(times))
                elif re.search(r"timeout|\u043d\u0435\u0434\u043e\u0441\u0442\u0438\u0436\u0438\u043c|timed out", line, re.IGNORECASE):
                    times.append(0)
                    self._bridge.chart_update.emit(list(times))
            self._proc.wait()
            valid = [t for t in times if t > 0]
            if valid:
                avg = sum(valid) / len(valid)
                self._bridge.finished.emit(success_count > 0,
                    f"Sent: {len(times)}  Received: {success_count}  Avg: {avg:.0f} ms")
            else:
                self._bridge.finished.emit(False, "Все запросы истекли по таймауту")
        except Exception as e:
            self._bridge.finished.emit(False, str(e))

    def refresh_theme(self, dark: bool) -> None:
        super().refresh_theme(dark)


class _TraceroutePanel(_ToolPanel):
    def __init__(self, dark: bool = True, i18n=None, runner=None) -> None:
        super().__init__("Traceroute", dark, i18n=i18n, runner=runner)
        self._host = QLineEdit()
        self._host.setObjectName("ToolInput")
        self._host.setPlaceholderText(self._tr("tools_placeholder_host"))
        self._host.setFixedHeight(28)
        self._host.returnPressed.connect(self._on_run)
        self._form.addWidget(self._host, 1)

    def retranslate(self) -> None:
        self._host.setPlaceholderText(self._tr("tools_placeholder_host"))

    def _on_run(self) -> None:
        host = self._host.text().strip()
        if not host:
            self._status.setText("Введите хост")
            return
        self._output.clear()
        self._set_running(True)
        self._status.setText(f"Tracing route to {host}...")
        threading.Thread(target=self._worker, args=(host,), daemon=True).start()

    def _worker(self, host: str) -> None:
        try:
            cmd = ["tracert", "-d", "-w", "2000", host]
            self._proc = self._runner.popen(cmd, encoding="cp866", errors="replace")
            hops = 0
            for line in self._proc.stdout or []:  # type: ignore[union-attr]
                if not self._running:
                    break
                line = line.rstrip()
                self._bridge.output.emit(line, False)
                if re.match(r"\s*\d+\s", line):
                    hops += 1
            self._proc.wait()
            self._bridge.finished.emit(True, f"Completed: {hops} hops")
        except Exception as e:
            self._bridge.finished.emit(False, str(e))


class _DnsPanel(_ToolPanel):
    def __init__(self, dark: bool = True, i18n=None, runner=None) -> None:
        super().__init__("DNS Lookup", dark, i18n=i18n, runner=runner)
        self._host = QLineEdit()
        self._host.setObjectName("ToolInput")
        self._host.setPlaceholderText(self._tr("tools_placeholder_domain"))
        self._host.setFixedHeight(28)
        self._host.returnPressed.connect(self._on_run)
        self._form.addWidget(self._host, 1)

        lbl = QLabel("Type:")
        lbl.setObjectName("FieldLabel")
        self._type = QComboBox()
        self._type.setObjectName("ToolCombo")
        self._type.addItems(["A", "AAAA", "MX", "NS", "PTR", "TXT", "SOA", "CNAME", "SRV"])
        self._type.setFixedWidth(90)
        self._type.setFixedHeight(28)
        self._form.addWidget(lbl)
        self._form.addWidget(self._type)

    def retranslate(self) -> None:
        self._host.setPlaceholderText(self._tr("tools_placeholder_domain"))

    def _on_run(self) -> None:
        host = self._host.text().strip()
        if not host:
            self._status.setText("Введите домен или IP")
            return
        self._output.clear()
        self._set_running(True)
        self._status.setText(f"Looking up {host}...")
        _ALLOWED_QTYPES = {"A", "AAAA", "MX", "NS", "PTR", "TXT", "SOA", "CNAME", "SRV"}
        qtype = self._type.currentText()
        if qtype not in _ALLOWED_QTYPES:
            qtype = "A"
        threading.Thread(target=self._worker, args=(host, qtype), daemon=True).start()

    @staticmethod
    def _decode_line(raw: bytes) -> str:
        # Try UTF-8 first (modern Windows with UTF-8 locale)
        try:
            return raw.decode("utf-8").rstrip()
        except UnicodeDecodeError:
            pass
        # CP1251 uses 0xC0–0xDF for Cyrillic uppercase (А–Я)
        # CP866  uses 0x80–0x9F for Cyrillic uppercase (А–Я)
        # Count bytes in each range to pick the right encoding
        cp1251_score = sum(1 for b in raw if 0xC0 <= b <= 0xDF)
        cp866_score  = sum(1 for b in raw if 0x80 <= b <= 0x9F)
        enc = "cp1251" if cp1251_score >= cp866_score else "cp866"
        return raw.decode(enc, errors="replace").rstrip()

    def _worker(self, host: str, qtype: str) -> None:
        try:
            cmd = ["nslookup", f"-type={qtype}", host]
            self._proc = self._runner.popen(cmd)
            lines = []
            for raw_line in iter(self._proc.stdout.readline, b""):  # type: ignore[union-attr]
                line = self._decode_line(raw_line)
                lines.append(line)
                self._bridge.output.emit(line, False)
            self._proc.wait()
            ok = any("Address" in ln or "Name" in ln or "\u0410\u0434\u0440\u0435\u0441" in ln for ln in lines)
            self._bridge.finished.emit(ok, "Готово" if ok else "Записи не найдены")
        except Exception as e:
            self._bridge.finished.emit(False, str(e))


class _FlushDnsPanel(_ToolPanel):
    def __init__(self, dark: bool = True, runner=None) -> None:
        super().__init__("Flush DNS Cache", dark, runner=runner)
        info = QLabel("Clears the Windows DNS resolver cache (ipconfig /flushdns)")
        info.setObjectName("FieldLabel")
        info.setWordWrap(True)
        self._form.addWidget(info, 1)
        self.btn_run.setText("\u25b6  Очистить")
        self.btn_stop.hide()

    def _on_run(self) -> None:
        self._output.clear()
        self._set_running(True)
        self._status.setText("Очистка DNS-кэша...")
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self) -> None:
        try:
            result = self._runner.run(["ipconfig", "/flushdns"], timeout=15)
            for line in (result.stdout + result.stderr).splitlines():
                self._bridge.output.emit(line, False)
            self._bridge.finished.emit(result.success, "DNS-кэш очищен" if result.success else "Ошибка")
        except Exception as e:
            self._bridge.finished.emit(False, str(e))


_IPCONFIG_PS = """\
[Console]::OutputEncoding = [System.Text.UTF8Encoding]::new($false)

function prefix2mask($n) {
    if ($n -lt 0 -or $n -gt 32) { return "" }
    $m = [uint32]0
    for ($i=0; $i -lt $n; $i++) { $m = $m -bor ([uint32]1 -shl (31-$i)) }
    return "{0}.{1}.{2}.{3}" -f ([byte](($m -shr 24) -band 255)),([byte](($m -shr 16) -band 255)),([byte](($m -shr 8) -band 255)),([byte]($m -band 255))
}

$allAdapters = Get-NetAdapter | Sort-Object InterfaceIndex
$allIPs      = Get-NetIPAddress      -EA SilentlyContinue
$allIfs      = Get-NetIPInterface    -EA SilentlyContinue
$allRoutes   = Get-NetRoute          -EA SilentlyContinue | Where-Object { $_.DestinationPrefix -in @('0.0.0.0/0','::0/0') }
$allDns      = Get-DnsClientServerAddress -EA SilentlyContinue

foreach ($a in $allAdapters) {
    $idx = $a.ifIndex
    $if4 = $allIfs | Where-Object { $_.InterfaceIndex -eq $idx -and $_.AddressFamily -eq 'IPv4' } | Select-Object -First 1
    $if6 = $allIfs | Where-Object { $_.InterfaceIndex -eq $idx -and $_.AddressFamily -eq 'IPv6' } | Select-Object -First 1

    $medium = "$($a.NdisPhysicalMediumType)"
    $ifType = if ($medium -match 'NativeWifi|WirelessLan|802[.]11') { "Wireless" } `
              elseif ($medium -match '802[.]3|Ethernet') { "Ethernet" } `
              elseif ($a.InterfaceType -eq 71) { "Wireless" } `
              elseif ($a.InterfaceType -eq 6)  { "Ethernet" } `
              else { $medium }

    Write-Output "ADAPTER_START:$($a.Name)"
    Write-Output "Description=$($a.InterfaceDescription)"
    Write-Output "MAC=$($a.MacAddress)"
    Write-Output "Interface Type=$ifType"
    Write-Output "Enabled=$($a.AdminStatus -eq 'Up')"
    Write-Output "Connected=$($a.MediaConnectionState -eq 'Connected')"
    Write-Output "Speed=$($a.LinkSpeed)"
    Write-Output "Interface Index=$idx"
    Write-Output "MTU=$(if ($if4) { $if4.NlMtu } else { '' })"

    # IPv4
    $ips4 = @($allIPs | Where-Object { $_.InterfaceIndex -eq $idx -and $_.AddressFamily -eq 'IPv4' })
    Write-Output "IPv4 - Enabled=$(if ($if4) { 'True' } else { 'False' })"
    Write-Output "IPv4 - DHCP=$(if ($if4) { $if4.Dhcp } else { '' })"
    foreach ($ip in $ips4) {
        Write-Output "IPv4 - IP=$($ip.IPAddress) : $(prefix2mask([int]$ip.PrefixLength))"
    }
    $gw4 = ($allRoutes | Where-Object { $_.InterfaceIndex -eq $idx -and $_.DestinationPrefix -eq '0.0.0.0/0' } | Select-Object -First 1).NextHop
    Write-Output "IPv4 - Gateway=$(if ($gw4) { $gw4 } else { '' })"
    $dns4 = $allDns | Where-Object { $_.InterfaceIndex -eq $idx -and $_.AddressFamily -eq 2 } | Select-Object -First 1
    Write-Output "IPv4 - DNS=$(if ($dns4 -and $dns4.ServerAddresses) { $dns4.ServerAddresses -join ', ' } else { '' })"
    Write-Output "IPv4 - Metric=$(if ($if4) { $if4.InterfaceMetric } else { '' })"

    # IPv6
    $ips6 = @($allIPs | Where-Object { $_.InterfaceIndex -eq $idx -and $_.AddressFamily -eq 'IPv6' })
    Write-Output "IPv6 - Enabled=$(if ($if6) { 'True' } else { 'False' })"
    Write-Output "IPv6 - DHCP=$(if ($if6) { $if6.Dhcp } else { '' })"
    $v6 = ($ips6 | ForEach-Object { "$($_.IPAddress)/$($_.PrefixLength)" }) -join '; '
    Write-Output "IPv6 - IP=$v6"
    $gw6 = ($allRoutes | Where-Object { $_.InterfaceIndex -eq $idx -and $_.DestinationPrefix -eq '::0/0' } | Select-Object -First 1).NextHop
    Write-Output "IPv6 - Gateway=$(if ($gw6) { $gw6 } else { '' })"
    $dns6 = $allDns | Where-Object { $_.InterfaceIndex -eq $idx -and $_.AddressFamily -eq 23 } | Select-Object -First 1
    Write-Output "IPv6 - DNS=$(if ($dns6 -and $dns6.ServerAddresses) { $dns6.ServerAddresses -join ', ' } else { '' })"
    Write-Output "IPv6 - Metric=$(if ($if6) { $if6.InterfaceMetric } else { '' })"

    Write-Output "ADAPTER_END"
}
"""


def _parse_ipconfig(text: str) -> list:
    """Парсит вывод PS-скрипта в список (name, [(key, val), ...])."""
    adapters, current_name, current_props = [], None, []
    for line in text.splitlines():
        line = line.strip()
        if line.startswith("ADAPTER_START:"):
            current_name = line[len("ADAPTER_START:"):]
            current_props = []
        elif line == "ADAPTER_END" and current_name is not None:
            adapters.append((current_name, current_props))
            current_name = None
        elif current_name is not None and "=" in line:
            key, _, val = line.partition("=")
            current_props.append((key.strip(), val.strip()))
    return adapters


class _IpconfigBridge(QObject):
    done       = Signal(list)
    status     = Signal(str)
    enable_btn = Signal()


class _IpconfigPanel(QWidget):
    def __init__(self, dark: bool = True, parent=None, i18n=None, runner=None) -> None:
        super().__init__(parent)
        self._dark = dark
        self._i18n = i18n
        self._runner = runner
        self._bridge = _IpconfigBridge()
        _t = lambda k: i18n.get(k) if i18n else k  # noqa: E731

        root = QVBoxLayout(self)
        root.setContentsMargins(16, 14, 16, 14)
        root.setSpacing(10)

        self._hdr = QLabel(_t("tools_ipconfig_title"))
        self._hdr.setObjectName("ToolPanelTitle")
        root.addWidget(self._hdr)

        _f = QFont("Segoe UI", 10)
        _f.setWeight(QFont.Weight.DemiBold)
        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        btn_row.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        self.btn_run = QPushButton("\u25b6  Обновить")
        self.btn_run.setProperty("role", "primary")
        self.btn_run.setObjectName("ToolBtn")
        self.btn_run.setFixedSize(100, 28)
        self.btn_run.setFont(_f)
        btn_row.addWidget(self.btn_run)
        btn_row.addStretch(1)
        root.addLayout(btn_row)

        self._tree = _CopyableTree()
        self._tree.setObjectName("IpconfigTree")
        self._tree.setColumnCount(2)
        self._tree.setHeaderLabels([_t("tools_ipconfig_col_param"), _t("tools_ipconfig_col_value")])
        self._tree.header().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self._tree.header().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._tree.header().setDefaultSectionSize(220)
        self._tree.setAlternatingRowColors(True)
        self._tree.setRootIsDecorated(True)
        self._tree.setIndentation(20)
        self._tree.setIconSize(QSize(16, 16))
        # Fusion style ensures QSS ::item:hover overrides work reliably on Windows
        self._tree.setStyle(QStyleFactory.create("Fusion"))
        root.addWidget(self._tree, 1)

        self._status = QLabel("")
        self._status.setObjectName("ToolStatus")
        root.addWidget(self._status)

        self._bridge.done.connect(self._populate)
        self._bridge.status.connect(self._status.setText)
        self._bridge.enable_btn.connect(lambda: self.btn_run.setEnabled(True))
        self.btn_run.clicked.connect(self._on_run)

    def _tr(self, key: str) -> str:
        return self._i18n.get(key) if self._i18n else key

    def retranslate(self) -> None:
        self._hdr.setText(self._tr("tools_ipconfig_title"))
        self._tree.setHeaderLabels([
            self._tr("tools_ipconfig_col_param"),
            self._tr("tools_ipconfig_col_value"),
        ])

    def _on_run(self) -> None:
        self._tree.clear()
        self.btn_run.setEnabled(False)
        self._status.setText("Загрузка адаптеров...")
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self) -> None:
        try:
            import base64
            encoded = base64.b64encode(_IPCONFIG_PS.encode('utf-16-le')).decode('ascii')
            result = self._runner.run(
                ["powershell", "-NoProfile", "-NonInteractive", "-EncodedCommand", encoded],
                timeout=30,
            )
            text = result.stdout or result.stderr
            adapters = _parse_ipconfig(text)
            self._bridge.done.emit(adapters)
            self._bridge.status.emit(f"Done — {len(adapters)} adapter(s)")
        except Exception as e:
            self._bridge.status.emit(str(e))
        finally:
            self._bridge.enable_btn.emit()

    def _populate(self, adapters: list) -> None:
        assets = Path(__file__).parent.parent / "assets"
        wifi_icon = QIcon(str(assets / "adapter-wifi.svg"))
        net_icon  = QIcon(str(assets / "adapter-network.svg"))

        self._tree.setUpdatesEnabled(False)
        self._tree.clear()
        bold = QFont("Segoe UI", 9)
        bold.setWeight(QFont.Weight.DemiBold)
        small = QFont("Segoe UI", 8)
        tops = []
        for name, props in adapters:
            n_lower = name.lower()
            top = QTreeWidgetItem([name, ""])
            top.setFont(0, bold)
            if any(w in n_lower for w in ("wi-fi", "wireless", "wlan", "wifi", "беспроводная")):
                top.setIcon(0, wifi_icon)
            else:
                top.setIcon(0, net_icon)
            for key, val in props:
                child = QTreeWidgetItem(top, [key, val])
                child.setFont(0, small)
                child.setFont(1, small)
            tops.append(top)
        self._tree.addTopLevelItems(tops)
        self._tree.setUpdatesEnabled(True)
        self._tree.resizeColumnToContents(0)

    def refresh_theme(self, dark: bool) -> None:
        self._dark = dark


class _PortScanPanel(_ToolPanel):
    _PRESET_KEYS = [
        ("tools_portscan_preset_basic",    "1-1024"),
        ("tools_portscan_preset_extended", "1-10000"),
        ("tools_portscan_preset_all",      "1-65535"),
        ("Web",                            "80, 443, 8080, 8443"),
        ("tools_portscan_preset_manual",   ""),
    ]

    def __init__(self, dark: bool = True, i18n=None) -> None:
        super().__init__("Port Scanner (TCP)", dark, i18n=i18n)
        self._host = QLineEdit()
        self._host.setObjectName("ToolInput")
        self._host.setPlaceholderText(self._tr("tools_placeholder_host"))
        self._host.setFixedHeight(28)
        self._host.returnPressed.connect(self._on_run)
        self._form.addWidget(self._host, 1)

        self._preset = QComboBox()
        self._preset.setObjectName("ToolCombo")
        self._preset.setFixedSize(150, 28)
        self._preset.addItems([
            i18n.get(k) if (i18n and not k.startswith("W")) else k
            for k, _ in self._PRESET_KEYS
        ])
        self._preset.currentIndexChanged.connect(self._on_preset_changed)
        self._form.addWidget(self._preset)

        self._ports = QLineEdit()
        self._ports.setObjectName("ToolInput")
        self._ports.setPlaceholderText("80, 443, 1000-2000")
        self._ports.setFixedSize(150, 28)
        self._ports.setText("1-1024")
        self._ports.returnPressed.connect(self._on_run)
        self._ports.textEdited.connect(self._on_ports_edited)
        self._form.addWidget(self._ports)

    def retranslate(self) -> None:
        self._host.setPlaceholderText(self._tr("tools_placeholder_host"))
        cur_idx = self._preset.currentIndex()
        self._preset.blockSignals(True)
        for i, (key, _) in enumerate(self._PRESET_KEYS):
            label = self._tr(key) if not key.startswith("W") else key
            self._preset.setItemText(i, label)
        self._preset.blockSignals(False)
        self._preset.setCurrentIndex(cur_idx)

    def _on_preset_changed(self, idx: int) -> None:
        spec = self._PRESET_KEYS[idx][1]
        if spec:
            self._ports.setText(spec)

    def _on_ports_edited(self) -> None:
        """При ручном редактировании поля портов — переключаем комбо на Custom."""
        custom_idx = len(self._PRESET_KEYS) - 1
        if self._preset.currentIndex() != custom_idx:
            self._preset.blockSignals(True)
            self._preset.setCurrentIndex(custom_idx)
            self._preset.blockSignals(False)

    @staticmethod
    def _parse_ports(spec: str) -> list[int]:
        """Парсит строку портов: '80', '80,443', '1-1024', '80,443,1000-2000'."""
        ports: list[int] = []
        for part in spec.split(","):
            part = part.strip()
            if not part:
                continue
            if "-" in part:
                lo, _, hi = part.partition("-")
                lo_i, hi_i = int(lo.strip()), int(hi.strip())
                ports.extend(range(lo_i, hi_i + 1))
            else:
                ports.append(int(part))
        return sorted(set(p for p in ports if 1 <= p <= 65535))

    def _on_run(self) -> None:
        host = self._host.text().strip()
        if not host:
            self._status.setText("Введите хост")
            return
        try:
            ports = self._parse_ports(self._ports.text())
        except ValueError:
            self._status.setText("Неверный формат портов — пример: 80, 443, 1000-2000")
            return
        if not ports:
            self._status.setText("Не указаны допустимые порты")
            return
        self._output.clear()
        self._set_running(True)
        self._status.setText(f"Scanning {host} — {len(ports)} port(s)...")
        threading.Thread(target=self._worker, args=(host, ports), daemon=True).start()

    def _worker(self, host: str, ports: list[int]) -> None:
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def probe(port: int) -> tuple[int, str]:
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.settimeout(0.5)
                    if s.connect_ex((host, port)) == 0:
                        try:
                            svc = socket.getservbyport(port)
                        except Exception:
                            svc = ""
                        return port, svc
            except Exception:
                pass
            return port, None  # type: ignore[return-value]

        try:
            total = len(ports)
            open_count = 0
            done_count = 0

            with ThreadPoolExecutor(max_workers=256) as ex:
                futures = {ex.submit(probe, p): p for p in ports}
                for fut in as_completed(futures):
                    if not self._running:
                        ex.shutdown(wait=False, cancel_futures=True)
                        break
                    port, svc = fut.result()
                    if svc is not None:
                        open_count += 1
                        self._bridge.output.emit(
                            f"  {port:5d}/tcp  OPEN  {svc}", False
                        )
                    done_count += 1
                    if done_count % 256 == 0 or done_count == total:
                        pct = min(100, int(done_count / total * 100))
                        self._bridge.output.emit(
                            f"  [{pct}%] Scanned {done_count}/{total} ports", False
                        )

            self._bridge.finished.emit(True, f"Scan complete: {open_count} open port(s)")
        except Exception as e:
            self._bridge.finished.emit(False, str(e))


class _NetstatBridge(QObject):
    rows_ready = Signal(list)
    finished   = Signal(bool, str)


class _NetstatPanel(QWidget):
    def __init__(self, dark: bool = True, i18n=None, runner=None) -> None:
        super().__init__()
        self._dark = dark
        self._i18n = i18n
        self._runner = runner
        self._bridge = _NetstatBridge()
        self._bridge.rows_ready.connect(self._populate)
        self._bridge.finished.connect(self._on_finished)
        _t = lambda k: i18n.get(k) if i18n else k  # noqa: E731

        root = QVBoxLayout(self)
        root.setContentsMargins(16, 14, 16, 14)
        root.setSpacing(10)

        self._hdr = QLabel(_t("tools_netstat_title"))
        self._hdr.setObjectName("ToolPanelTitle")
        root.addWidget(self._hdr)

        _f = QFont("Segoe UI", 10)
        _f.setWeight(QFont.Weight.DemiBold)
        form = QHBoxLayout()
        form.setSpacing(8)
        self._filter = QComboBox()
        self._filter.setObjectName("ToolCombo")
        self._filter.setFixedHeight(28)
        self._filter.addItems([_t("tools_filter_all"), "TCP", "UDP", "LISTENING", "ESTABLISHED"])
        form.addWidget(self._filter, 0, Qt.AlignmentFlag.AlignVCenter)
        form.addStretch(1)
        root.addLayout(form)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        self.btn_run = QPushButton("▶  Запуск")
        self.btn_run.setProperty("role", "primary")
        self.btn_run.setObjectName("ToolBtn")
        self.btn_run.setFixedSize(90, 28)
        self.btn_run.setFont(_f)
        self.btn_run.clicked.connect(self._on_run)
        self.btn_stop = QPushButton("■  Стоп")
        self.btn_stop.setProperty("role", "action")
        self.btn_stop.setObjectName("ToolBtn")
        self.btn_stop.setFixedSize(80, 28)
        self.btn_stop.setFont(_f)
        self.btn_stop.setEnabled(False)
        self.btn_clear = QPushButton("Очистить")
        self.btn_clear.setProperty("role", "action")
        self.btn_clear.setObjectName("ToolBtn")
        self.btn_clear.setFixedSize(90, 28)
        self.btn_clear.setFont(_f)
        self.btn_clear.clicked.connect(self._clear)
        btn_row.addWidget(self.btn_run, 0, Qt.AlignmentFlag.AlignVCenter)
        btn_row.addWidget(self.btn_stop, 0, Qt.AlignmentFlag.AlignVCenter)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_clear, 0, Qt.AlignmentFlag.AlignVCenter)
        root.addLayout(btn_row)

        self._table = _CopyableTree()
        self._table.setObjectName("NetstatTable")
        self._table.setStyle(QStyleFactory.create("Fusion"))
        self._table.setStyleSheet(_TREE_SS_LIGHT if not dark else _TREE_SS_DARK)
        self._table.setRootIsDecorated(False)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setColumnCount(5)
        self._table.setHeaderLabels(["Протокол", "Локальный", "Удалённый", "Состояние", "PID"])
        self._table.header().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.header().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._table.header().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self._table.header().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self._table.header().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        root.addWidget(self._table, 1)

        self._status = QLabel("")
        self._status.setObjectName("ToolStatus")
        root.addWidget(self._status)

    def _clear(self) -> None:
        self._table.clear()
        self._status.setText("")

    def _on_run(self) -> None:
        self._table.clear()
        self.btn_run.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self._running = True
        self._status.setText("Получение соединений...")
        threading.Thread(target=self._worker, daemon=True).start()

    def _on_finished(self, ok: bool, msg: str) -> None:
        self.btn_run.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self._running = False
        color = "#10B981" if ok else "#EF4444"
        self._status.setStyleSheet(f"color: {color}; font-size: 12px;")
        self._status.setText(msg)

    def _populate(self, rows: list) -> None:
        self._table.setUpdatesEnabled(False)
        for proto, local, remote, state, pid in rows:
            item = QTreeWidgetItem([proto, local, remote, state, pid])
            item.setTextAlignment(4, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self._table.addTopLevelItem(item)
        self._table.setUpdatesEnabled(True)

    def _worker(self) -> None:
        try:
            result = self._runner.run(["netstat", "-ano"], timeout=15)
            flt = self._filter.currentText()
            rows = []
            for line in result.stdout.splitlines():
                parts = line.split()
                if len(parts) < 4:
                    continue
                proto = parts[0].upper()
                if proto not in ("TCP", "UDP"):
                    continue
                if len(parts) == 5:
                    proto, local, remote, state, pid = parts
                elif len(parts) == 4:
                    proto, local, remote, pid = parts
                    state = "—"
                else:
                    continue
                if self._filter.currentIndex() != 0:  # 0 = All
                    if flt in ("TCP", "UDP") and proto != flt:
                        continue
                    if flt in ("LISTENING", "ESTABLISHED") and state.upper() != flt:
                        continue
                rows.append((proto, local, remote, state, pid))
            self._bridge.rows_ready.emit(rows)
            self._bridge.finished.emit(True, f"Показано строк: {len(rows)}")
        except Exception as e:
            self._bridge.finished.emit(False, str(e))

    def _tr(self, key: str) -> str:
        return self._i18n.get(key) if self._i18n else key

    def retranslate(self) -> None:
        self._hdr.setText(self._tr("tools_netstat_title"))
        cur_idx = self._filter.currentIndex()
        self._filter.blockSignals(True)
        self._filter.setItemText(0, self._tr("tools_filter_all"))
        self._filter.blockSignals(False)
        self._filter.setCurrentIndex(cur_idx)
        self._table.setHeaderLabels([
            self._tr("tools_netstat_col_proto"),
            self._tr("tools_netstat_col_local"),
            self._tr("tools_netstat_col_remote"),
            self._tr("tools_netstat_col_state"),
            "PID",
        ])

    def refresh_theme(self, dark: bool) -> None:
        self._dark = dark
        self._table.setStyleSheet(_TREE_SS_LIGHT if not dark else _TREE_SS_DARK)


class _ArpBridge(QObject):
    rows_ready = Signal(list)   # list of (iface, ip, mac, type)
    finished   = Signal(bool, str)


class _ArpPanel(QWidget):
    def __init__(self, dark: bool = True, i18n=None, runner=None) -> None:
        super().__init__()
        self._dark = dark
        self._i18n = i18n
        self._runner = runner
        self._bridge = _ArpBridge()
        self._bridge.rows_ready.connect(self._on_rows_ready)
        self._bridge.finished.connect(self._on_finished)
        self._build()

    def _tr(self, key: str) -> str:
        return self._i18n.get(key) if self._i18n else key

    def _build(self) -> None:
        _f = QFont()
        _f.setWeight(QFont.Weight.DemiBold)
        root = QVBoxLayout(self)
        root.setContentsMargins(18, 16, 18, 14)
        root.setSpacing(8)

        self._hdr = QLabel(self._tr("tools_arp_title"))
        self._hdr.setObjectName("ToolPanelTitle")
        root.addWidget(self._hdr)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        self.btn_run = QPushButton(self._tr("tools_btn_run"))
        self.btn_run.setProperty("role", "primary")
        self.btn_run.setObjectName("ToolBtn")
        self.btn_run.setFixedSize(90, 28)
        self.btn_run.setFont(_f)
        self.btn_run.clicked.connect(self._on_run)
        self.btn_clear = QPushButton(self._tr("tools_btn_clear"))
        self.btn_clear.setProperty("role", "action")
        self.btn_clear.setObjectName("ToolBtn")
        self.btn_clear.setFixedSize(90, 28)
        self.btn_clear.setFont(_f)
        self.btn_clear.clicked.connect(self._on_clear)
        btn_row.addWidget(self.btn_run)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_clear)
        root.addLayout(btn_row)

        self._table = _CopyableTree()
        self._table.setObjectName("NetstatTable")
        self._table.setStyle(QStyleFactory.create("Fusion"))
        self._table.setStyleSheet(_TREE_SS_LIGHT if not self._dark else _TREE_SS_DARK)
        self._table.setRootIsDecorated(False)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setColumnCount(4)
        self._table.setHeaderLabels(["IP-адрес", "MAC-адрес", "Тип", "Интерфейс"])
        self._table.header().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.header().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self._table.header().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self._table.header().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        root.addWidget(self._table, 1)

        self._status = QLabel("")
        self._status.setObjectName("ToolStatus")
        root.addWidget(self._status)

    def retranslate(self) -> None:
        self._hdr.setText(self._tr("tools_arp_title"))
        self.btn_run.setText(self._tr("tools_btn_run"))
        self.btn_clear.setText(self._tr("tools_btn_clear"))

    def _on_run(self) -> None:
        self._table.setRowCount(0) if hasattr(self._table, "setRowCount") else None
        self.btn_run.setEnabled(False)
        self._status.setText("")
        threading.Thread(target=self._worker, daemon=True).start()

    def _on_clear(self) -> None:
        self._table.clear()
        self._status.setText("")

    def _worker(self) -> None:
        try:
            result = self._runner.run(["arp", "-a"], timeout=10)
            rows = []
            iface = ""
            for line in result.stdout.splitlines():
                # Interface header: "Интерфейс: 192.168.1.1 --- 0x3"
                m_iface = re.match(
                    r"^\s*(?:Interface|Интерфейс)\s*[:：]\s*([\d.]+)", line, re.IGNORECASE
                )
                if m_iface:
                    iface = m_iface.group(1)
                    continue
                # Data row: "  192.168.1.1   aa-bb-cc-dd-ee-ff   динамический"
                m_row = re.match(
                    r"^\s*([\d.]+)\s+([\da-fA-F]{2}[-:][\da-fA-F]{2}[-:][\da-fA-F]{2}"
                    r"[-:][\da-fA-F]{2}[-:][\da-fA-F]{2}[-:][\da-fA-F]{2})\s+(\S+)",
                    line,
                )
                if m_row:
                    rows.append((iface, m_row.group(1), m_row.group(2), m_row.group(3)))
            self._bridge.rows_ready.emit(rows)
            self._bridge.finished.emit(True, f"Записей в ARP: {len(rows)}")
        except Exception as exc:
            self._bridge.finished.emit(False, str(exc))

    def _on_rows_ready(self, rows: list) -> None:
        self._table.clear()
        for iface, ip, mac, kind in rows:
            item = QTreeWidgetItem([ip, mac, kind, iface])
            item.setFont(0, QFont("Consolas", 9))
            item.setFont(1, QFont("Consolas", 9))
            self._table.addTopLevelItem(item)
        self._table.resizeColumnToContents(0)
        self._table.resizeColumnToContents(1)
        self._table.resizeColumnToContents(2)

    def _on_finished(self, success: bool, msg: str) -> None:
        self.btn_run.setEnabled(True)
        color = "#22C55E" if success else "#EF4444"
        self._status.setText(msg)
        self._status.setStyleSheet(f"color: {color}; font-size: 12px;")

    def refresh_theme(self, dark: bool) -> None:
        self._dark = dark
        self._table.setStyleSheet(_TREE_SS_LIGHT if not dark else _TREE_SS_DARK)


class _HttpCheckPanel(_ToolPanel):
    def __init__(self, dark: bool = True) -> None:
        super().__init__("HTTP Check", dark)
        self._url = QLineEdit()
        self._url.setObjectName("ToolInput")
        self._url.setPlaceholderText("https://example.com")
        self._url.setFixedHeight(28)
        self._url.returnPressed.connect(self._on_run)
        self._form.addWidget(self._url, 1)

    def _on_run(self) -> None:
        url = self._url.text().strip()
        if not url:
            self._status.setText("Введите URL")
            return
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
            self._url.setText(url)
        self._output.clear()
        self._set_running(True)
        self._status.setText(f"Проверка {url}...")
        threading.Thread(target=self._worker, args=(url,), daemon=True).start()

    def _worker(self, url: str) -> None:
        try:
            import urllib.request
            import urllib.error
            import time
            import ssl

            ctx = ssl.create_default_context()  # TLS verification enabled by default

            _CONNECT_TIMEOUT = 8   # seconds to establish connection
            _READ_TIMEOUT    = 10  # seconds to read response headers
            _MAX_BODY_BYTES  = 4 * 1024  # read at most 4 KB of body

            self._bridge.output.emit(f"  URL:      {url}", False)
            self._bridge.output.emit("  TLS verify: ON", False)

            redirects = []
            current = url
            for _ in range(10):
                req = urllib.request.Request(current, headers={"User-Agent": "NetConneXion/1.0"})
                t0 = time.monotonic()
                try:
                    with urllib.request.urlopen(
                        req, timeout=_READ_TIMEOUT, context=ctx
                    ) as resp:
                        resp.read(_MAX_BODY_BYTES)  # drain limited body to complete handshake
                        elapsed = int((time.monotonic() - t0) * 1000)
                        status = resp.status
                        final_url = resp.url
                        headers = dict(resp.headers)
                        content_type = headers.get("Content-Type", "—")
                        content_len = headers.get("Content-Length", "—")
                        server = headers.get("Server", "—")

                        if redirects:
                            self._bridge.output.emit(f"\n  Редиректы ({len(redirects)}):", False)
                            for r in redirects:
                                self._bridge.output.emit(f"    → {r}", False)

                        self._bridge.output.emit(f"\n  Статус:       {status}", False)
                        self._bridge.output.emit(f"  Время:        {elapsed} ms", False)
                        self._bridge.output.emit(f"  Финальный URL:{final_url}", False)
                        self._bridge.output.emit(f"  Content-Type: {content_type}", False)
                        self._bridge.output.emit(f"  Content-Len:  {content_len}", False)
                        self._bridge.output.emit(f"  Server:       {server}", False)

                        # TLS info
                        if hasattr(resp, "fp") and hasattr(resp.fp, "raw"):
                            sock = getattr(resp.fp.raw, "_sock", None)
                            if sock and hasattr(sock, "cipher"):
                                cipher = sock.cipher()
                                self._bridge.output.emit(f"  TLS:          {cipher[1]} / {cipher[0]}", False)

                        self._bridge.finished.emit(True, f"HTTP {status} — {elapsed} ms")
                        return
                except urllib.error.HTTPError as e:
                    elapsed = int((time.monotonic() - t0) * 1000)
                    self._bridge.output.emit(f"\n  Статус: {e.code} {e.reason}", True)
                    self._bridge.finished.emit(False, f"HTTP {e.code} — {elapsed} ms")
                    return
                except urllib.error.URLError as e:
                    if hasattr(e, "reason") and "Moved" in str(e.reason):
                        redirects.append(current)
                        current = str(e.reason)
                        continue
                    raise
            self._bridge.finished.emit(False, "Слишком много редиректов")
        except Exception as e:
            self._bridge.finished.emit(False, str(e))


class _SslPanel(_ToolPanel):
    def __init__(self, dark: bool = True, i18n=None) -> None:
        super().__init__("SSL Certificate", dark, i18n=i18n)
        self._host = QLineEdit()
        self._host.setObjectName("ToolInput")
        self._host.setPlaceholderText(self._tr("tools_placeholder_ssl"))
        self._host.setFixedHeight(28)
        self._host.returnPressed.connect(self._on_run)
        self._form.addWidget(self._host, 1)

    def retranslate(self) -> None:
        self._host.setPlaceholderText(self._tr("tools_placeholder_ssl"))

    def _on_run(self) -> None:
        host = self._host.text().strip().removeprefix("https://").removeprefix("http://").rstrip("/")
        if not host:
            self._status.setText("Введите хост")
            return
        self._output.clear()
        self._set_running(True)
        self._status.setText(f"Получение сертификата {host}...")
        threading.Thread(target=self._worker, args=(host,), daemon=True).start()

    def _worker(self, host: str) -> None:
        import ssl
        import socket
        import datetime
        try:
            if ":" in host:
                hostname, port_s = host.rsplit(":", 1)
                port = int(port_s)
            else:
                hostname, port = host, 443

            ctx = ssl.create_default_context()
            with socket.create_connection((hostname, port), timeout=8) as sock:
                with ctx.wrap_socket(sock, server_hostname=hostname) as ssock:
                    cert = ssock.getpeercert()
                    cipher = ssock.cipher()

            if cert is None:
                self._bridge.finished.emit(False, "Сертификат не получен")
                return

            def fmt_name(fields) -> str:
                return ", ".join(f"{k}={v}" for rdn in (fields or []) for k, v in rdn)

            subject    = fmt_name(cert.get("subject"))
            issuer     = fmt_name(cert.get("issuer"))
            not_before = str(cert.get("notBefore", "—"))
            not_after  = str(cert.get("notAfter",  "—"))

            days_left: int | None = None
            try:
                exp = datetime.datetime.strptime(not_after, "%b %d %H:%M:%S %Y %Z")
                days_left = (exp - datetime.datetime.utcnow()).days
                expiry_str = f"{not_after}  ({days_left} дн.)"
            except Exception:
                expiry_str = not_after

            san_raw  = cert.get("subjectAltName") or []
            san_list = [str(v) for t, v in san_raw if t == "DNS"]

            tls_ver = cipher[1] if cipher else "—"
            tls_alg = cipher[0] if cipher else "—"

            lines = [
                f"  {'Хост':<18} {hostname}:{port}",
                f"  {'Subject':<18} {subject}",
                f"  {'Issuer':<18} {issuer}",
                f"  {'Valid From':<18} {not_before}",
                f"  {'Valid To':<18} {expiry_str}",
                f"  {'TLS версия':<18} {tls_ver}",
                f"  {'Шифр':<18} {tls_alg}",
                "",
                f"  SAN ({len(san_list)}):",
            ]
            for s in san_list:
                lines.append(f"    • {s}")

            for line in lines:
                self._bridge.output.emit(line, False)

            if days_left is None:
                self._bridge.finished.emit(True, "Сертификат получен")
            elif days_left < 0:
                self._bridge.finished.emit(False, f"Сертификат истёк {-days_left} дн. назад")
            elif days_left <= 30:
                self._bridge.finished.emit(False, f"Истекает через {days_left} дн. — требует обновления")
            else:
                self._bridge.finished.emit(True, f"Действителен ещё {days_left} дн.")
        except ssl.SSLCertVerificationError as e:
            self._bridge.finished.emit(False, f"Ошибка верификации: {e}")
        except Exception as e:
            self._bridge.finished.emit(False, str(e))


class _RouteTableBridge(QObject):
    rows_ready = Signal(list)
    finished   = Signal(bool, str)


class _RouteTablePanel(QWidget):
    def __init__(self, dark: bool = True, i18n=None, runner=None) -> None:
        super().__init__()
        self._dark = dark
        self._i18n = i18n
        self._runner = runner
        self._bridge = _RouteTableBridge()
        self._bridge.rows_ready.connect(self._populate)
        self._bridge.finished.connect(self._on_finished)
        _t = lambda k: i18n.get(k) if i18n else k  # noqa: E731

        root = QVBoxLayout(self)
        root.setContentsMargins(16, 14, 16, 14)
        root.setSpacing(10)

        self._hdr = QLabel(_t("tools_route_title"))
        self._hdr.setObjectName("ToolPanelTitle")
        root.addWidget(self._hdr)

        _f = QFont("Segoe UI", 10)
        _f.setWeight(QFont.Weight.DemiBold)
        form = QHBoxLayout()
        form.setSpacing(8)
        self._filter = QComboBox()
        self._filter.setObjectName("ToolCombo")
        self._filter.setFixedHeight(28)
        self._filter.addItems([_t("tools_route_filter_both"), _t("tools_route_filter_ipv4"), _t("tools_route_filter_ipv6")])
        form.addWidget(self._filter, 0, Qt.AlignmentFlag.AlignVCenter)
        form.addStretch(1)
        root.addLayout(form)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        self.btn_run = QPushButton("▶  Запуск")
        self.btn_run.setProperty("role", "primary")
        self.btn_run.setObjectName("ToolBtn")
        self.btn_run.setFixedSize(90, 28)
        self.btn_run.setFont(_f)
        self.btn_run.clicked.connect(self._on_run)
        self.btn_clear = QPushButton("Очистить")
        self.btn_clear.setProperty("role", "action")
        self.btn_clear.setObjectName("ToolBtn")
        self.btn_clear.setFixedSize(90, 28)
        self.btn_clear.setFont(_f)
        self.btn_clear.clicked.connect(self._clear)
        btn_row.addWidget(self.btn_run, 0, Qt.AlignmentFlag.AlignVCenter)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_clear, 0, Qt.AlignmentFlag.AlignVCenter)
        root.addLayout(btn_row)

        self._table = _CopyableTree()
        self._table.setObjectName("NetstatTable")
        self._table.setStyle(QStyleFactory.create("Fusion"))
        self._table.setStyleSheet(_TREE_SS_LIGHT if not dark else _TREE_SS_DARK)
        self._table.setRootIsDecorated(False)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setColumnCount(5)
        self._table.setHeaderLabels([_t("tools_route_col_net"), _t("tools_route_col_mask"), _t("tools_route_col_gw"), _t("tools_route_col_iface"), _t("tools_route_col_metric")])
        hdr = self._table.header()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        root.addWidget(self._table, 1)

        self._status = QLabel("")
        self._status.setObjectName("ToolStatus")
        root.addWidget(self._status)

    def _clear(self) -> None:
        self._table.clear()
        self._status.setText("")

    def _on_run(self) -> None:
        self._table.clear()
        self.btn_run.setEnabled(False)
        self._status.setText("Получение маршрутов...")
        threading.Thread(target=self._worker, daemon=True).start()

    def _on_finished(self, ok: bool, msg: str) -> None:
        self.btn_run.setEnabled(True)
        color = "#10B981" if ok else "#EF4444"
        self._status.setStyleSheet(f"color: {color}; font-size: 12px;")
        self._status.setText(msg)

    def _populate(self, rows: list) -> None:
        self._table.setUpdatesEnabled(False)
        for row in rows:
            item = QTreeWidgetItem(row)
            item.setTextAlignment(4, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self._table.addTopLevelItem(item)
        self._table.setUpdatesEnabled(True)

    @staticmethod
    @staticmethod
    def _clean(s: str) -> str:
        """Оставляет только ASCII-печатные символы из имён адаптеров Windows."""
        return "".join(c for c in s if 0x20 <= ord(c) <= 0x7E).strip()

    def _tr(self, key: str) -> str:
        return self._i18n.get(key) if self._i18n else key

    def retranslate(self) -> None:
        self._hdr.setText(self._tr("tools_route_title"))
        cur_idx = self._filter.currentIndex()
        self._filter.blockSignals(True)
        self._filter.setItemText(0, self._tr("tools_route_filter_both"))
        self._filter.setItemText(1, self._tr("tools_route_filter_ipv4"))
        self._filter.setItemText(2, self._tr("tools_route_filter_ipv6"))
        self._filter.blockSignals(False)
        self._filter.setCurrentIndex(cur_idx)
        self._table.setHeaderLabels([
            self._tr("tools_route_col_net"),
            self._tr("tools_route_col_mask"),
            self._tr("tools_route_col_gw"),
            self._tr("tools_route_col_iface"),
            self._tr("tools_route_col_metric"),
        ])

    def _worker(self) -> None:
        try:
            flt_idx = self._filter.currentIndex()
            rows: list[list[str]] = []

            if flt_idx != 2:  # not IPv6-only
                # IPv4 маршруты через PowerShell — надёжный парсинг
                ps = (
                    "Get-NetRoute -AddressFamily IPv4 | "
                    "Select-Object DestinationPrefix,NextHop,InterfaceAlias,RouteMetric | "
                    "ConvertTo-Json -Compress"
                )
                enc = base64.b64encode(ps.encode("utf-16-le")).decode("ascii")
                result = self._runner.run(
                    ["powershell", "-NonInteractive", "-EncodedCommand", enc],
                    timeout=15,
                )
                if result.stdout.strip():
                    data = json.loads(result.stdout)
                    if isinstance(data, dict):
                        data = [data]
                    for r in data:
                        dest    = r.get("DestinationPrefix", "")
                        gw      = r.get("NextHop", "") or "—"
                        iface   = self._clean(r.get("InterfaceAlias", ""))
                        metric  = str(r.get("RouteMetric", ""))
                        if "/" in dest:
                            net, prefix = dest.split("/", 1)
                        else:
                            net, prefix = dest, ""
                        rows.append([net, f"/{prefix}", gw, iface, metric])

            if flt_idx != 1:  # not IPv4-only
                ps6 = (
                    "Get-NetRoute -AddressFamily IPv6 | "
                    "Select-Object DestinationPrefix,NextHop,InterfaceAlias,RouteMetric | "
                    "ConvertTo-Json -Compress"
                )
                enc6 = base64.b64encode(ps6.encode("utf-16-le")).decode("ascii")
                r6 = self._runner.run(
                    ["powershell", "-NonInteractive", "-EncodedCommand", enc6],
                    timeout=15,
                )
                if r6.stdout.strip():
                    data6 = json.loads(r6.stdout)
                    if isinstance(data6, dict):
                        data6 = [data6]
                    for r in data6:
                        dest   = r.get("DestinationPrefix", "")
                        gw     = r.get("NextHop", "") or "—"
                        iface  = self._clean(r.get("InterfaceAlias", ""))
                        metric = str(r.get("RouteMetric", ""))
                        if "/" in dest:
                            net, prefix = dest.split("/", 1)
                        else:
                            net, prefix = dest, ""
                        rows.append([net, f"/{prefix}", gw, iface, metric])

            self._bridge.rows_ready.emit(rows)
            self._bridge.finished.emit(True, f"Маршрутов: {len(rows)}")
        except Exception as e:
            self._bridge.finished.emit(False, str(e))

    def refresh_theme(self, dark: bool) -> None:
        self._dark = dark
        self._table.setStyleSheet(_TREE_SS_LIGHT if not dark else _TREE_SS_DARK)


class _SignalGraph(QWidget):
    """Живой график уровня сигнала Wi-Fi (dBm)."""
    MAX_POINTS = 60
    LEVELS = [(-50, "#10B981"), (-65, "#F59E0B"), (-75, "#EF4444")]

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._values: list[float] = []
        self._dark = True
        self.setMinimumHeight(90)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

    def push(self, dbm: float) -> None:
        self._values.append(dbm)
        if len(self._values) > self.MAX_POINTS:
            self._values.pop(0)
        self.update()

    def set_dark(self, dark: bool) -> None:
        self._dark = dark
        self.update()

    def paintEvent(self, _) -> None:  # noqa: N802
        from PySide6.QtCore import QPointF
        p = QPainter(self)
        try:
            p.setRenderHint(QPainter.RenderHint.Antialiasing)
            w, h = self.width(), self.height()

            bg = QColor("#1E293B") if self._dark else QColor("#FFFFFF")
            p.fillRect(0, 0, w, h, bg)

            dbm_min, dbm_max = -100.0, -40.0

            def y_of(dbm: float) -> float:
                return h - (dbm - dbm_min) / (dbm_max - dbm_min) * h

            grid_color = QColor(255, 255, 255, 25) if self._dark else QColor(0, 0, 0, 20)
            text_color = QColor(148, 163, 184) if self._dark else QColor(100, 116, 139)
            for lvl in (-50, -60, -70, -80, -90):
                y = y_of(lvl)
                p.setPen(QPen(grid_color, 1, Qt.PenStyle.DashLine))
                p.drawLine(0, int(y), w, int(y))
                p.setPen(text_color)
                p.setFont(QFont("Segoe UI", 8))
                p.drawText(4, int(y) - 2, f"{lvl}")

            if not self._values:
                return

            last = self._values[-1]
            if last >= -65:
                line_color = QColor("#10B981")
            elif last >= -75:
                line_color = QColor("#F59E0B")
            else:
                line_color = QColor("#EF4444")

            step = w / max(self.MAX_POINTS - 1, 1)
            offset = (self.MAX_POINTS - len(self._values)) * step
            points = [
                QPointF(offset + i * step, y_of(v))
                for i, v in enumerate(self._values)
            ]
            pen = QPen(line_color, 2)
            pen.setCapStyle(Qt.PenCapStyle.RoundCap)
            pen.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
            p.setPen(pen)
            for i in range(len(points) - 1):
                p.drawLine(points[i], points[i + 1])

            p.setBrush(line_color)
            p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(points[-1], 4, 4)
        finally:
            p.end()


class _SignalMonitorBridge(QObject):
    updated = Signal(dict)   # данные опроса
    roam    = Signal(dict)   # событие роуминга
    stopped = Signal()


class _SignalMonitorPanel(QWidget):

    # Порог «слабый сигнал» для учащения опроса
    _WEAK_DBM = -70.0

    def __init__(self, dark: bool = True, i18n=None, runner=None) -> None:
        super().__init__()
        self._dark = dark
        self._i18n = i18n
        self._runner = runner
        self._running = False
        self._bridge = _SignalMonitorBridge()
        self._bridge.updated.connect(self._on_update)
        self._bridge.roam.connect(self._on_roam)
        self._bridge.stopped.connect(self._on_stopped)

        root = QVBoxLayout(self)
        root.setContentsMargins(16, 14, 16, 14)
        root.setSpacing(10)

        _hdr = QLabel("Wi-Fi Signal Monitor")
        _hdr.setObjectName("ToolPanelTitle")
        root.addWidget(_hdr)

        # ── Кнопки ───────────────────────────────────────────────────
        _f = QFont("Segoe UI", 10)
        _f.setWeight(QFont.Weight.DemiBold)
        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        btn_row.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        _t = lambda k: i18n.get(k) if i18n else k  # noqa: E731
        self._btn_start = QPushButton(_t("tools_signal_btn_start"))
        self._btn_start.setProperty("role", "primary")
        self._btn_start.setObjectName("ToolBtn")
        self._btn_start.setFixedSize(90, 28)
        self._btn_start.setFont(_f)
        self._btn_start.clicked.connect(self._start)
        self._btn_stop = QPushButton(_t("tools_signal_btn_stop"))
        self._btn_stop.setProperty("role", "action")
        self._btn_stop.setObjectName("ToolBtn")
        self._btn_stop.setFixedSize(80, 28)
        self._btn_stop.setFont(_f)
        self._btn_stop.setEnabled(False)
        self._btn_stop.clicked.connect(self._stop)
        btn_row.addWidget(self._btn_start)
        btn_row.addWidget(self._btn_stop)
        btn_row.addStretch(1)
        root.addLayout(btn_row)

        # ── Карточки текущих значений ─────────────────────────────────
        cards = QHBoxLayout()
        cards.setSpacing(10)
        self._lbl_dbm    = self._card("dBm", "—", cards)
        self._lbl_pct    = self._card("%",   "—", cards)
        self._lbl_ssid   = self._card("SSID",  "—", cards, stretch=2)
        self._lbl_bssid  = self._card("BSSID", "—", cards, stretch=2)
        self._lbl_ch     = self._card("CH",   "—", cards)
        self._lbl_band   = self._card("Band", "—", cards)
        self._lbl_rx     = self._card("Rx Mbps", "—", cards)
        self._lbl_tx     = self._card("Tx Mbps", "—", cards)
        root.addLayout(cards)

        # ── График ───────────────────────────────────────────────────
        self._graph = _SignalGraph()
        self._graph.set_dark(dark)
        root.addWidget(self._graph)

        # ── Лог роуминга ──────────────────────────────────────────────
        log_lbl = QLabel("Roaming Log")
        log_lbl.setObjectName("ToolPanelTitle")
        log_lbl.setStyleSheet("font-size: 12px;")
        root.addWidget(log_lbl)
        self._log = QTextEdit()
        self._log.setObjectName("ToolOutput")
        self._log.setReadOnly(True)
        self._log.setFont(QFont("Consolas", 9))
        self._log.setMaximumHeight(160)
        root.addWidget(self._log)

        self._status = QLabel("")
        self._status.setObjectName("ToolStatus")
        root.addWidget(self._status)

    # ── Вспомогательный метод карточки ───────────────────────────────

    def _card(self, label: str, value: str, layout: QHBoxLayout,
              stretch: int = 1) -> QLabel:
        frame = QFrame()
        frame.setObjectName("SignalCard")
        bg = "rgba(255,255,255,0.06)" if self._dark else "#EEF2FF"
        frame.setStyleSheet(
            f"QFrame#SignalCard{{background:{bg};border-radius:8px;}}"
            f"QFrame#SignalCard QLabel{{background:transparent;}}"
        )
        col = QVBoxLayout(frame)
        col.setContentsMargins(10, 8, 10, 8)
        col.setSpacing(3)
        lbl_key = QLabel(label)
        key_color = "#94A3B8" if self._dark else "#6366F1"
        lbl_key.setStyleSheet(f"color:{key_color};font-size:10px;font-weight:600;letter-spacing:0.5px;")
        val_color = "#F1F5F9" if self._dark else "#1E293B"
        val_lbl = QLabel(value)
        val_lbl.setObjectName("SignalCardValue")
        val_lbl.setStyleSheet(f"font-size:17px;font-weight:700;color:{val_color};")
        col.addWidget(lbl_key)
        col.addWidget(val_lbl)
        layout.addWidget(frame, stretch)
        return val_lbl

    def _tr(self, key: str) -> str:
        return self._i18n.get(key) if self._i18n else key

    def retranslate(self) -> None:
        self._btn_start.setText(self._tr("tools_signal_btn_start"))
        self._btn_stop.setText(self._tr("tools_signal_btn_stop"))

    # ── Управление ────────────────────────────────────────────────────

    def _start(self) -> None:
        self._running = True
        self._btn_start.setEnabled(False)
        self._btn_stop.setEnabled(True)
        self._log.clear()
        import threading as _t
        _t.Thread(target=self._poll_loop, daemon=True).start()

    def _stop(self) -> None:
        self._running = False

    def _on_stopped(self) -> None:
        self._btn_start.setEnabled(True)
        self._btn_stop.setEnabled(False)
        self._status.setText("Остановлено")

    # ── Обновление UI ─────────────────────────────────────────────────

    def _on_update(self, d: dict) -> None:
        dbm = d.get("dbm", 0.0)
        pct = d.get("signal", 0)
        if dbm <= -75:
            dbm_color = "#EF4444"
        elif dbm <= -65:
            dbm_color = "#F59E0B"
        else:
            dbm_color = "#10B981"

        val_color = "#F1F5F9" if self._dark else "#1E293B"
        val_style = f"font-size:17px;font-weight:700;color:{val_color};"

        self._lbl_dbm.setText(f"{dbm:.0f}")
        self._lbl_dbm.setStyleSheet(f"font-size:17px;font-weight:700;color:{dbm_color};")
        self._lbl_pct.setText(f"{pct}")
        self._lbl_pct.setStyleSheet(val_style)
        self._lbl_ssid.setText(d.get("ssid", "—"))
        self._lbl_ssid.setStyleSheet(val_style)
        self._lbl_bssid.setText(d.get("bssid", "—"))
        self._lbl_bssid.setStyleSheet(val_style)
        self._lbl_ch.setText(str(d.get("channel", "—")))
        self._lbl_ch.setStyleSheet(val_style)
        self._lbl_band.setText(d.get("band", "—"))
        self._lbl_band.setStyleSheet(val_style)
        self._lbl_rx.setText(str(d.get("rx", "—")))
        self._lbl_rx.setStyleSheet(val_style)
        self._lbl_tx.setText(str(d.get("tx", "—")))
        self._lbl_tx.setStyleSheet(val_style)
        self._graph.push(dbm)
        self._status.setStyleSheet("color:#64748B;font-size:11px;")
        interval = "1" if dbm < self._WEAK_DBM else "2"
        self._status.setText(
            self._tr("tools_signal_status").format(
                interval=interval, count=len(self._graph._values)
            )
        )

    def _on_roam(self, d: dict) -> None:
        import datetime
        ts = datetime.datetime.now().strftime("%H:%M:%S.%f")[:-3]
        old_b = d.get("old_bssid", "?")
        new_b = d.get("new_bssid", "?")
        old_dbm = d.get("old_dbm", 0.0)
        new_dbm = d.get("new_dbm", 0.0)
        ms = d.get("ms", "?")
        self._log.append(
            f"[{ts}]  Роуминг  {old_b}  →  {new_b}\n"
            f"         Сигнал до: {old_dbm:.0f} dBm  |  после: {new_dbm:.0f} dBm  |  ~{ms} ms"
        )
        self._log.append("")

    # ── Фоновый поток ─────────────────────────────────────────────────

    @staticmethod
    def _parse(output: str) -> dict:
        """Парсит вывод netsh wlan show interfaces построчно (EN + RU)."""
        # Строим словарь: нормализованный_ключ -> значение
        kv: dict[str, str] = {}
        for line in output.splitlines():
            if ":" not in line:
                continue
            idx = line.index(":")
            key = line[:idx].strip().lower()
            val = line[idx + 1:].strip()
            if key and val:
                kv[key] = val

        def v(*keys: str) -> str:
            for k in keys:
                if k in kv:
                    return kv[k]
            return ""

        state = v("state", "состояние")
        connected = any(w in state.lower() for w in ("connect", "подключ"))

        # SSID — ключ "ssid", но не "bssid"
        ssid = ""
        for line in output.splitlines():
            stripped = line.strip()
            if re.match(r"^SSID\s*:", stripped, re.IGNORECASE) and not re.match(r"^BSSID", stripped, re.IGNORECASE):
                ssid = stripped.split(":", 1)[1].strip()
                break

        signal_raw = v("signal", "сигнал").replace("%", "").strip()
        pct = int(signal_raw) if signal_raw.isdigit() else 0
        dbm = (pct / 2.0) - 100.0

        rx_raw = v("receive rate (mbps)", "скорость получения (мбит/с)",
                   "скорость получени", "скорость приема")
        tx_raw = v("transmit rate (mbps)", "скорость передачи (мбит/с)", "скорость передачи")

        radio = v("radio type", "тип радиосвязи", "тип сети")
        channel_raw = v("channel", "канал")
        try:
            ch_num = int(channel_raw)
        except (ValueError, TypeError):
            ch_num = 0
        band_raw = v("диапазон", "band", "frequency band")
        if "5" in band_raw or ch_num > 14:
            band = "5 GHz"
        else:
            band = "2.4 GHz"

        # BSSID — ищем и "BSSID" и "AP BSSID"
        import re as _re
        bssid_m = _re.search(r"(?:AP\s+)?BSSID\s*:\s*([0-9a-fA-F]{2}(?:[:\-][0-9a-fA-F]{2}){5})",
                             output, _re.IGNORECASE)
        bssid = bssid_m.group(1) if bssid_m else ""

        return {
            "ssid":      ssid,
            "bssid":     bssid,
            "signal":    pct,
            "dbm":       dbm,
            "channel":   channel_raw,
            "band":      band,
            "radio":     radio,
            "rx":        rx_raw.split(".")[0] if rx_raw else "—",
            "tx":        tx_raw.split(".")[0] if tx_raw else "—",
            "state":     state,
            "connected": connected,
        }

    def _poll_loop(self) -> None:
        import time
        last_bssid: str = ""
        last_dbm: float = -100.0
        roam_start_ts: float = 0.0

        while self._running:
            try:
                result = self._runner.run(
                    ["netsh", "wlan", "show", "interfaces"], timeout=5,
                )
                d = self._parse(result.stdout)

                if not d.get("connected", False):
                    self._bridge.updated.emit({
                        "dbm": -100.0, "signal": 0,
                        "ssid": "Нет подключения",
                        "bssid": "—", "channel": "—", "band": "—", "rx": "—", "tx": "—",
                        "connected": False,
                    })
                else:
                    bssid = d.get("bssid", "")
                    dbm   = d.get("dbm", -100.0)

                    # Начало потенциального роуминга
                    if dbm < self._WEAK_DBM and roam_start_ts == 0.0:
                        roam_start_ts = time.monotonic()

                    # Событие роуминга
                    if last_bssid and bssid and bssid != last_bssid:
                        ms = int((time.monotonic() - roam_start_ts) * 1000) if roam_start_ts else "?"
                        self._bridge.roam.emit({
                            "old_bssid": last_bssid,
                            "new_bssid": bssid,
                            "old_dbm":   last_dbm,
                            "new_dbm":   dbm,
                            "ms":        ms,
                        })
                        roam_start_ts = 0.0

                    last_bssid = bssid
                    last_dbm   = dbm
                    self._bridge.updated.emit(d)

                    # Учащаем опрос при слабом сигнале, но не чаще 1с
                    # чтобы не мешать WLAN AutoConfig сервису Windows
                    interval = 1.0 if dbm < self._WEAK_DBM else 2.0
                    time.sleep(interval)
                    continue

            except Exception as e:
                self._bridge.updated.emit({
                    "dbm": -100.0, "signal": 0, "ssid": f"Ошибка: {e}",
                    "bssid": "—", "channel": "—", "band": "—", "rx": "—", "tx": "—",
                })

            time.sleep(2.0)

        self._bridge.stopped.emit()

    def refresh_theme(self, dark: bool) -> None:
        self._dark = dark
        self._graph.set_dark(dark)


class _SubnetCalcPanel(QWidget):
    _ROW_KEYS = [
        ("tools_subnet_cidr",      "cidr"),
        ("tools_subnet_network",   "network"),
        ("tools_subnet_mask",      "mask"),
        ("tools_subnet_wildcard",  "wildcard"),
        ("tools_subnet_broadcast", "broadcast"),
        ("tools_subnet_first",     "first"),
        ("tools_subnet_last",      "last"),
        ("tools_subnet_hosts",     "hosts"),
        ("tools_subnet_cls",       "cls"),
    ]

    def __init__(self, dark: bool = True, i18n=None) -> None:
        super().__init__()
        self._dark = dark
        self._i18n = i18n
        _t = lambda k: i18n.get(k) if i18n else k  # noqa: E731
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 14, 16, 14)
        root.setSpacing(10)

        hdr = QLabel("Subnet Calculator")
        hdr.setObjectName("ToolPanelTitle")
        root.addWidget(hdr)

        form = QHBoxLayout()
        form.setSpacing(8)
        self._cidr = QLineEdit()
        self._cidr.setObjectName("ToolInput")
        self._cidr.setPlaceholderText(_t("tools_subnet_placeholder"))
        self._cidr.setFixedHeight(28)
        self._cidr.returnPressed.connect(self._calc)
        form.addWidget(self._cidr, 1)
        _f = QFont("Segoe UI", 10)
        _f.setWeight(QFont.Weight.DemiBold)
        self._btn_calc = QPushButton(_t("tools_subnet_btn_calc"))
        self._btn_calc.setProperty("role", "primary")
        self._btn_calc.setObjectName("ToolBtn")
        self._btn_calc.setFixedHeight(28)
        self._btn_calc.setFont(_f)
        self._btn_calc.clicked.connect(self._calc)
        form.addWidget(self._btn_calc)
        root.addLayout(form)

        self._grid_w = QWidget()
        gl = QGridLayout(self._grid_w)
        gl.setSpacing(6)
        gl.setContentsMargins(0, 4, 0, 4)
        gl.setColumnStretch(1, 1)
        self._fields: dict = {}
        self._row_labels: dict = {}
        for row, (i18n_key, field_key) in enumerate(self._ROW_KEYS):
            lbl = QLabel(_t(i18n_key) + ":")
            lbl.setObjectName("SubnetLabel")
            val = QLabel("\u2014")
            val.setObjectName("SubnetValue")
            val.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            gl.addWidget(lbl, row, 0)
            gl.addWidget(val, row, 1)
            self._fields[field_key] = val
            self._row_labels[field_key] = lbl
        self._grid_w.setVisible(False)
        root.addWidget(self._grid_w)

        self._status = QLabel("")
        self._status.setObjectName("ToolStatus")
        root.addWidget(self._status)
        root.addStretch(1)

    def _tr(self, key: str) -> str:
        return self._i18n.get(key) if self._i18n else key

    def retranslate(self) -> None:
        self._btn_calc.setText(self._tr("tools_subnet_btn_calc"))
        self._cidr.setPlaceholderText(self._tr("tools_subnet_placeholder"))
        for i18n_key, field_key in self._ROW_KEYS:
            self._row_labels[field_key].setText(self._tr(i18n_key) + ":")

    def _calc(self) -> None:
        import ipaddress
        text = self._cidr.text().strip()
        if not text:
            self._status.setText("Введите адрес")
            return
        try:
            if "/" in text:
                ip_part, mask_part = text.split("/", 1)
                if "." in mask_part:
                    prefix = bin(int(ipaddress.IPv4Address(mask_part))).count("1")
                    text = f"{ip_part}/{prefix}"
            net = ipaddress.IPv4Network(text, strict=False)
            first_octet = int(str(net.network_address).split(".")[0])
            if first_octet < 128:
                cls = "A"
            elif first_octet < 192:
                cls = "B"
            elif first_octet < 224:
                cls = "C"
            elif first_octet < 240:
                cls = "D (Multicast)"
            else:
                cls = "E (Reserved)"
            hosts = max(0, net.num_addresses - 2) if net.prefixlen < 31 else net.num_addresses
            first = str(net.network_address + 1) if net.prefixlen < 31 else str(net.network_address)
            last  = str(net.broadcast_address - 1) if net.prefixlen < 31 else str(net.broadcast_address)
            self._fields["cidr"].setText(str(net))
            self._fields["network"].setText(str(net.network_address))
            self._fields["mask"].setText(str(net.netmask))
            self._fields["wildcard"].setText(str(net.hostmask))
            self._fields["broadcast"].setText(str(net.broadcast_address))
            self._fields["first"].setText(first)
            self._fields["last"].setText(last)
            self._fields["hosts"].setText(f"{hosts:,}".replace(",", "\u202f"))
            self._fields["cls"].setText(cls)
            self._grid_w.setVisible(True)
            self._status.setText(f"/{net.prefixlen} — {hosts} хостов")
            self._status.setStyleSheet("color: #22C55E; font-size: 12px;")
        except Exception as e:
            self._status.setText(f"Ошибка: {e}")
            self._status.setStyleSheet("color: #EF4444; font-size: 12px;")
            self._grid_w.setVisible(False)

    def refresh_theme(self, dark: bool) -> None:
        self._dark = dark


class _DnsCacheBridge(QObject):
    rows_ready = Signal(list)
    finished   = Signal(bool, str)


class _DnsCachePanel(QWidget):
    def __init__(self, dark: bool = True, i18n=None, runner=None) -> None:
        super().__init__()
        self._dark = dark
        self._i18n = i18n
        self._runner = runner
        self._bridge = _DnsCacheBridge()

        root = QVBoxLayout(self)
        root.setContentsMargins(16, 14, 16, 14)
        root.setSpacing(10)

        hdr = QLabel("DNS Cache")
        hdr.setObjectName("ToolPanelTitle")
        root.addWidget(hdr)

        _f = QFont("Segoe UI", 10)
        _f.setWeight(QFont.Weight.DemiBold)
        _t = lambda k: i18n.get(k) if i18n else k  # noqa: E731
        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        self._btn_ref = QPushButton(_t("tools_dns_cache_btn_refresh"))
        self._btn_ref.setProperty("role", "primary")
        self._btn_ref.setObjectName("ToolBtn")
        self._btn_ref.setFixedHeight(28)
        self._btn_ref.setFont(_f)
        self._btn_flush = QPushButton(_t("tools_dns_cache_btn_flush"))
        self._btn_flush.setProperty("role", "action")
        self._btn_flush.setObjectName("ToolBtn")
        self._btn_flush.setFixedHeight(28)
        self._btn_flush.setFont(_f)
        btn_row.addWidget(self._btn_ref, 0, Qt.AlignmentFlag.AlignVCenter)
        btn_row.addWidget(self._btn_flush, 0, Qt.AlignmentFlag.AlignVCenter)
        btn_row.addStretch(1)
        root.addLayout(btn_row)

        self._tree = _CopyableTree()
        self._tree.setObjectName("ToolTree")
        self._tree.setStyle(QStyleFactory.create("Fusion"))
        self._tree.setStyleSheet(_TREE_SS_LIGHT if not dark else _TREE_SS_DARK)
        self._tree.setHeaderLabels([_t("tools_dns_cache_col_name"), _t("tools_dns_cache_col_type"), "TTL", _t("tools_dns_cache_col_data")])
        self._tree.header().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._tree.header().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self._tree.header().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self._tree.header().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self._tree.setAlternatingRowColors(True)
        self._tree.setRootIsDecorated(False)
        self._tree.setSortingEnabled(True)
        root.addWidget(self._tree, 1)

        self._status = QLabel("")
        self._status.setObjectName("ToolStatus")
        root.addWidget(self._status)

        self._btn_ref.clicked.connect(self._refresh)
        self._btn_flush.clicked.connect(self._flush)
        self._bridge.rows_ready.connect(self._populate)
        self._bridge.finished.connect(self._on_finished)

    def _tr(self, key: str) -> str:
        return self._i18n.get(key) if self._i18n else key

    def retranslate(self) -> None:
        self._btn_ref.setText(self._tr("tools_dns_cache_btn_refresh"))
        self._btn_flush.setText(self._tr("tools_dns_cache_btn_flush"))
        self._tree.setHeaderLabels([
            self._tr("tools_dns_cache_col_name"),
            self._tr("tools_dns_cache_col_type"),
            "TTL",
            self._tr("tools_dns_cache_col_data"),
        ])

    def _refresh(self) -> None:
        self._tree.clear()
        self._status.setText("Получение DNS кэша...")
        threading.Thread(target=self._worker, daemon=True).start()

    def _flush(self) -> None:
        reply = QMessageBox.question(
            self, self._tr("dlg_dns_flush_title"),
            self._tr("dlg_dns_flush_text"),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            self._runner.run(["ipconfig", "/flushdns"], timeout=5)
            self._tree.clear()
            self._status.setText("DNS кэш очищен")
            self._status.setStyleSheet("color: #22C55E; font-size: 12px;")
        except Exception as e:
            self._status.setText(str(e))
            self._status.setStyleSheet("color: #EF4444; font-size: 12px;")

    def _worker(self) -> None:
        try:
            result = self._runner.run(["ipconfig", "/displaydns"], timeout=15)
            rows = self._parse(result.stdout)
            self._bridge.rows_ready.emit(rows)
            self._bridge.finished.emit(True, f"Записей: {len(rows)}")
        except Exception as e:
            self._bridge.finished.emit(False, str(e))

    # Поля которые содержат имя записи (EN + RU)
    _NAME_KEYS  = {"Record Name", "Имя записи"}
    _TYPE_KEYS  = {"Record Type", "Тип записи"}
    _TTL_KEYS   = {"Time To Live", "Срок жизни"}
    _SKIP_KEYS  = {"Data Length", "Длина данных", "Section", "Раздел"}

    @staticmethod
    def _parse(text: str) -> list:
        rows: list = []
        name = ""
        rec_type = ""
        ttl = ""
        entry: dict = {}

        def _flush() -> None:
            nonlocal name, rec_type, ttl, entry
            if entry and name:
                data = next(iter(entry.values()), "")
                if data:
                    rows.append((name, rec_type, ttl, data))
            rec_type = ttl = ""
            entry = {}

        for line in text.splitlines():
            s = line.strip()
            if not s or set(s) <= {"-"}:
                _flush()
                continue
            if " : " in s:
                key, _, val = s.partition(" : ")
                key = key.rstrip(". ").strip()
                val = val.strip()
                if key in _DnsCachePanel._NAME_KEYS:
                    _flush()
                    name = val.rstrip(".")
                elif key in _DnsCachePanel._TYPE_KEYS:
                    rec_type = val
                elif key in _DnsCachePanel._TTL_KEYS:
                    ttl = val
                elif key not in _DnsCachePanel._SKIP_KEYS:
                    entry[key] = val
        _flush()
        return rows

    def _populate(self, rows: list) -> None:
        self._tree.clear()
        for rec_name, rec_type, ttl, data in rows:
            QTreeWidgetItem(self._tree, [rec_name, rec_type, ttl, data])

    def _on_finished(self, success: bool, msg: str) -> None:
        color = "#22C55E" if success else "#EF4444"
        self._status.setText(msg)
        self._status.setStyleSheet(f"color: {color}; font-size: 12px;")

    def refresh_theme(self, dark: bool) -> None:
        self._dark = dark
        self._tree.setStyleSheet(_TREE_SS_LIGHT if not dark else _TREE_SS_DARK)


# ── IP Batch Check ────────────────────────────────────────────────────────────

# Each worker spawns a ping.exe subprocess — keep under OS process limit.
# Windows starts degrading around 100+ simultaneous child processes.
_BATCH_MAX_WORKERS = 100
_BATCH_DEFAULT_WORKERS = 50


def _safe_workers(n_ips: int) -> int:
    """Return a safe worker count for *n_ips* addresses.

    Rules:
    - Never more than _BATCH_MAX_WORKERS (subprocess limit)
    - Never more than the number of IPs (no idle threads)
    - Scale down for small batches: no point in 50 threads for 10 IPs
    """
    return min(_BATCH_MAX_WORKERS, max(1, n_ips, min(n_ips, _BATCH_DEFAULT_WORKERS)))


class _IpBatchBridge(QObject):
    row_done   = Signal(int, str, str, str)   # row_idx, status, ms, hostname
    finished   = Signal(int, int)             # total, ok_count
    progress   = Signal(int, int, int)        # done, total, ok_count


def _ping_one(ip: str, timeout: int = 1, runner=None) -> tuple[bool, str]:
    """Ping *ip* once. Returns (reachable, rtt_ms_str)."""
    try:
        cmd = ["ping", "-n", "1", "-w", str(timeout * 1000), ip]
        r = runner.run(cmd, timeout=timeout + 2)
        if r.success:
            m = re.search(r"[Вв]ремя[<=](\d+)\s*мс|[Tt]ime[<=](\d+)\s*ms|[Tt]ime=(\d+)ms", r.stdout)
            ms = m.group(1) or m.group(2) or m.group(3) if m else "0"
            return True, ms
    except Exception:
        pass
    return False, ""


def _resolve_one(ip: str) -> str:
    """Reverse-DNS lookup for *ip*. Returns hostname or empty string."""
    try:
        return socket.gethostbyaddr(ip)[0]
    except Exception:
        return ""


class _IpBatchPanel(QWidget):
    """Batch ping + reverse-DNS for a list of IPs loaded from CSV/Excel."""

    def __init__(self, dark: bool = True, i18n=None, runner=None) -> None:
        super().__init__()
        self._dark = dark
        self._i18n = i18n
        self._runner = runner
        self._rows: list[dict] = []      # original rows from file
        self._headers: list[str] = []    # column names
        self._ip_col: str = ""           # selected IP column
        self._stop_flag = False
        self._done_count = 0
        self._bridge = _IpBatchBridge()
        self._bridge.row_done.connect(self._on_row_done)
        self._bridge.finished.connect(self._on_finished)
        self._bridge.progress.connect(self._on_progress)
        self._build()

    def _tr(self, key: str, **kwargs) -> str:
        s = self._i18n.get(key) if self._i18n else key
        return s.format(**kwargs) if kwargs else s

    # ── UI ────────────────────────────────────────────────────────

    @staticmethod
    def _step_label(n: int, text: str) -> QLabel:
        lbl = QLabel(f"  {n}. {text}")
        lbl.setStyleSheet("color: #94A3B8; font-size: 11px; font-weight: 600;")
        return lbl

    @staticmethod
    def _divider() -> QFrame:
        d = QFrame()
        d.setFrameShape(QFrame.Shape.HLine)
        d.setObjectName("ToolDivider")
        return d

    def _build(self) -> None:
        _f = QFont()
        _f.setWeight(QFont.Weight.DemiBold)

        root = QVBoxLayout(self)
        root.setContentsMargins(18, 14, 18, 14)
        root.setSpacing(0)

        hdr = QLabel(self._tr("tools_batch_title"))
        hdr.setObjectName("ToolPanelTitle")
        root.addWidget(hdr)
        root.addSpacing(10)

        # ── Step 1 ────────────────────────────────────────────────
        self._lbl_step1 = self._step_label(1, self._tr("tools_batch_step1"))
        root.addWidget(self._lbl_step1)
        root.addSpacing(5)

        row1 = QHBoxLayout()
        row1.setSpacing(8)
        self._btn_open = QPushButton(self._tr("tools_batch_open_btn"))
        self._btn_open.setProperty("role", "action")
        self._btn_open.setObjectName("ToolBtn")
        self._btn_open.setFixedSize(130, 28)
        self._btn_open.setFont(_f)
        self._btn_open.clicked.connect(self._on_open)
        row1.addWidget(self._btn_open)
        self._lbl_file = QLabel(self._tr("tools_batch_no_file"))
        self._lbl_file.setObjectName("ToolStatus")
        row1.addWidget(self._lbl_file, 1)
        root.addLayout(row1)
        root.addSpacing(10)
        root.addWidget(self._divider())
        root.addSpacing(10)

        # ── Step 2 ────────────────────────────────────────────────
        self._lbl_step2 = self._step_label(2, self._tr("tools_batch_step2"))
        root.addWidget(self._lbl_step2)
        root.addSpacing(5)

        row2 = QHBoxLayout()
        row2.setSpacing(8)
        self._lbl_col = QLabel(self._tr("tools_batch_col_label"))
        row2.addWidget(self._lbl_col)
        self._col_combo = QComboBox()
        self._col_combo.setObjectName("ToolCombo")
        self._col_combo.setMinimumWidth(200)
        self._col_combo.setEnabled(False)
        self._col_combo.currentTextChanged.connect(self._on_col_changed)
        row2.addWidget(self._col_combo)
        row2.addStretch(1)
        root.addLayout(row2)
        root.addSpacing(10)
        root.addWidget(self._divider())
        root.addSpacing(10)

        # ── Step 3 ────────────────────────────────────────────────
        self._lbl_step3 = self._step_label(3, self._tr("tools_batch_step3"))
        root.addWidget(self._lbl_step3)
        root.addSpacing(5)

        row3 = QHBoxLayout()
        row3.setSpacing(8)
        self._lbl_timeout = QLabel(self._tr("tools_batch_timeout_label"))
        row3.addWidget(self._lbl_timeout)
        self._ed_timeout = QLineEdit("1")
        self._ed_timeout.setObjectName("ToolInput")
        self._ed_timeout.setFixedWidth(44)
        self._ed_timeout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        row3.addWidget(self._ed_timeout)
        self._lbl_timeout_hint = QLabel(self._tr("tools_batch_timeout_hint"))
        self._lbl_timeout_hint.setStyleSheet("color: #64748B; font-size: 11px;")
        row3.addWidget(self._lbl_timeout_hint)
        row3.addSpacing(24)
        self._lbl_workers = QLabel(self._tr("tools_batch_workers_label"))
        row3.addWidget(self._lbl_workers)
        self._ed_workers = QLineEdit(str(_BATCH_DEFAULT_WORKERS))
        self._ed_workers.setObjectName("ToolInput")
        self._ed_workers.setFixedWidth(44)
        self._ed_workers.setAlignment(Qt.AlignmentFlag.AlignCenter)
        row3.addWidget(self._ed_workers)
        lbl_w = QLabel(f"(1–{_BATCH_MAX_WORKERS})")
        lbl_w.setStyleSheet("color: #64748B; font-size: 11px;")
        row3.addWidget(lbl_w)
        row3.addStretch(1)
        root.addLayout(row3)
        root.addSpacing(10)
        root.addWidget(self._divider())
        root.addSpacing(10)

        # ── Step 4 ────────────────────────────────────────────────
        self._lbl_step4 = self._step_label(4, self._tr("tools_batch_step4"))
        root.addWidget(self._lbl_step4)
        root.addSpacing(3)

        row4 = QHBoxLayout()
        row4.setSpacing(6)
        row4.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        self.btn_run = QPushButton(self._tr("tools_batch_run_btn"))
        self.btn_run.setProperty("role", "primary")
        self.btn_run.setObjectName("ToolBtn")
        self.btn_run.setFixedSize(90, 28)
        self.btn_run.setFont(_f)
        self.btn_run.setEnabled(False)
        self.btn_run.clicked.connect(self._on_run)
        self.btn_stop = QPushButton(self._tr("tools_batch_stop_btn"))
        self.btn_stop.setProperty("role", "action")
        self.btn_stop.setObjectName("ToolBtn")
        self.btn_stop.setFixedSize(80, 28)
        self.btn_stop.setFont(_f)
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self._on_stop)
        self._status = QLabel("")
        self._status.setObjectName("ToolStatus")
        row4.addWidget(self.btn_run)
        row4.addWidget(self.btn_stop)
        row4.addSpacing(12)
        row4.addWidget(self._status, 1)
        root.addLayout(row4)
        root.addSpacing(14)

        # Progress bar
        self._progress = QProgressBar()
        self._progress.setRange(0, 100)
        self._progress.setValue(0)
        self._progress.setFixedHeight(5)
        self._progress.setTextVisible(False)
        root.addWidget(self._progress)
        root.addSpacing(10)

        # ── Results table ─────────────────────────────────────────
        self._table = _CopyableTable(0, 0)
        self._table.setObjectName("NetstatTable")
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setDefaultSectionSize(22)
        root.addWidget(self._table, 1)

        # ── Save button ───────────────────────────────────────────
        row5 = QHBoxLayout()
        row5.setAlignment(Qt.AlignmentFlag.AlignRight)
        self._btn_export = QPushButton(self._tr("tools_batch_save_btn"))
        self._btn_export.setProperty("role", "action")
        self._btn_export.setObjectName("ToolBtn")
        self._btn_export.setFixedHeight(28)
        self._btn_export.setFont(_f)
        self._btn_export.setEnabled(False)
        self._btn_export.clicked.connect(self._on_export)
        row5.addWidget(self._btn_export)
        root.addSpacing(6)
        root.addLayout(row5)

    # ── File loading ─────────────────────────────────────────────

    def _on_open(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, self._tr("tools_batch_dlg_open"),
            "", self._tr("tools_batch_dlg_filter")
        )
        if not path:
            return
        try:
            self._rows, self._headers = self._load_file(path)
        except Exception as exc:
            self._status.setText(self._tr("tools_batch_err_read", err=exc))
            self._status.setStyleSheet("color: #EF4444; font-size: 12px;")
            return

        self._lbl_file.setText(self._tr("tools_batch_loaded", name=Path(path).name, n=len(self._rows)))
        self._col_combo.setEnabled(True)
        self._col_combo.blockSignals(True)
        self._col_combo.clear()
        self._col_combo.addItems(self._headers)
        # auto-select first column that looks like IPs
        auto = next(
            (h for h in self._headers if re.search(r"ip|addr|хост|host", h, re.IGNORECASE)),
            self._headers[0] if self._headers else "",
        )
        self._col_combo.setCurrentText(auto)
        self._col_combo.blockSignals(False)
        self._ip_col = self._col_combo.currentText()

        self._build_table(self._rows, self._headers, self._ip_col)
        self.btn_run.setEnabled(bool(self._rows))
        self._btn_export.setEnabled(False)
        self._progress.setValue(0)
        # Auto-tune worker count based on batch size
        self._ed_workers.setText(str(_safe_workers(len(self._rows))))
        self._status.setText(self._tr("tools_batch_status_loaded", n=len(self._rows)))
        self._status.setStyleSheet("color: #94A3B8; font-size: 12px;")

    @staticmethod
    def _load_file(path: str) -> tuple[list[dict], list[str]]:
        p = path.lower()
        if p.endswith(".csv"):
            # Try utf-8-sig first, fall back to cp1251 for Windows-exported files
            for enc in ("utf-8-sig", "cp1251", "utf-8"):
                try:
                    with open(path, newline="", encoding=enc) as f:
                        sample = f.read(4096)
                    # Auto-detect delimiter: count candidates in header line
                    first_line = sample.splitlines()[0] if sample else ""
                    dialect_delim = max(",", ";", "\t", "|",
                                        key=lambda d: first_line.count(d))
                    with open(path, newline="", encoding=enc) as f:
                        reader = csv.DictReader(f, delimiter=dialect_delim)
                        rows = list(reader)
                        headers = list(reader.fieldnames) if reader.fieldnames else (
                            list(rows[0].keys()) if rows else []
                        )
                    return rows, headers
                except UnicodeDecodeError:
                    continue
            raise ValueError("Не удалось определить кодировку CSV")
        # Excel (.xlsx / .xls)
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError("Установите openpyxl: pip install openpyxl") from exc
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return [], []
        headers = [str(c) if c is not None else f"col{i}" for i, c in enumerate(all_rows[0])]
        rows = [
            {headers[i]: (str(cell) if cell is not None else "") for i, cell in enumerate(row)}
            for row in all_rows[1:]
        ]
        return rows, headers

    def _on_col_changed(self, col: str) -> None:
        self._ip_col = col
        if self._rows:
            self._build_table(self._rows, self._headers, col)
            self._btn_export.setEnabled(False)

    def _build_table(self, rows: list[dict], headers: list[str], ip_col: str) -> None:
        """Rebuild results table: original columns + Status, RTT, Hostname."""
        extra = [
            self._tr("tools_batch_col_status"),
            self._tr("tools_batch_col_rtt"),
            self._tr("tools_batch_col_hostname"),
        ]
        all_cols = headers + extra
        self._table.setRowCount(0)
        self._table.setColumnCount(len(all_cols))
        self._table.setHorizontalHeaderLabels(all_cols)
        self._table.setRowCount(len(rows))
        for r_idx, row in enumerate(rows):
            for c_idx, col in enumerate(headers):
                item = QTableWidgetItem(str(row.get(col, "")))
                if col == ip_col:
                    item.setFont(QFont("Consolas", 9))
                self._table.setItem(r_idx, c_idx, item)
            for c_idx, _ in enumerate(extra):
                self._table.setItem(r_idx, len(headers) + c_idx, QTableWidgetItem(""))
        self._table.resizeColumnsToContents()
        self._table.horizontalHeader().setStretchLastSection(True)

    # ── Run / Stop ────────────────────────────────────────────────

    def _on_run(self) -> None:
        if not self._ip_col or not self._rows:
            return
        self._stop_flag = False
        self._done_count = 0
        self._btn_export.setEnabled(False)
        self.btn_run.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self._progress.setValue(0)
        # Reset result columns
        n_orig = len(self._headers)
        for r_idx in range(len(self._rows)):
            for c in range(3):
                self._table.setItem(r_idx, n_orig + c, QTableWidgetItem(""))
        try:
            timeout = max(1, min(10, int(self._ed_timeout.text())))
        except ValueError:
            timeout = 1
        self._ed_timeout.setText(str(timeout))
        try:
            workers = max(1, min(_BATCH_MAX_WORKERS, int(self._ed_workers.text())))
        except ValueError:
            workers = _BATCH_DEFAULT_WORKERS
        self._ed_workers.setText(str(workers))
        threading.Thread(
            target=self._worker,
            args=(list(self._rows), self._ip_col, self._headers, timeout, workers),
            daemon=True,
        ).start()

    def _on_stop(self) -> None:
        self._stop_flag = True

    def _worker(
        self, rows: list[dict], ip_col: str, headers: list[str], timeout: int, max_workers: int
    ) -> None:
        total = len(rows)
        ok_count = 0
        lock = threading.Lock()

        def _check(idx: int, row: dict):
            if self._stop_flag:
                return idx, "—", "", ""
            ip = str(row.get(ip_col, "")).strip()
            if not ip:
                return idx, "—", "", ""
            reachable, ms = _ping_one(ip, timeout, runner=self._runner)
            hostname = _resolve_one(ip) if reachable else ""
            status = "OK" if reachable else "Timeout"
            return idx, status, ms, hostname

        # Hard cap: never exceed _BATCH_MAX_WORKERS regardless of UI value,
        # and never create more threads than there are IPs to check.
        effective = min(max_workers, _BATCH_MAX_WORKERS, len(rows))
        with ThreadPoolExecutor(max_workers=effective) as pool:
            futures = {pool.submit(_check, i, row): i for i, row in enumerate(rows)}
            for fut in as_completed(futures):
                if self._stop_flag:
                    pool.shutdown(wait=False, cancel_futures=True)
                    break
                idx, status, ms, hostname = fut.result()
                with lock:
                    if status == "OK":
                        ok_count += 1
                    done = self._done_count + 1
                    self._done_count = done
                self._bridge.row_done.emit(idx, status, ms, hostname)
                self._bridge.progress.emit(done, total, ok_count)

        self._bridge.finished.emit(total, ok_count)

    # ── Signals ───────────────────────────────────────────────────

    def _on_row_done(self, row_idx: int, status: str, ms: str, hostname: str) -> None:
        n = len(self._headers)
        ok = status == "OK"
        color = "#22C55E" if ok else ("#94A3B8" if status == "—" else "#EF4444")
        for c_off, text in enumerate([status, ms, hostname]):
            item = QTableWidgetItem(text)
            if c_off == 0:
                item.setForeground(QColor(color))
                item.setFont(QFont("Consolas", 9))
            self._table.setItem(row_idx, n + c_off, item)

    def _on_progress(self, done: int, total: int, ok: int) -> None:
        pct = int(done / total * 100) if total else 0
        self._progress.setValue(pct)
        self._status.setText(self._tr("tools_batch_progress", done=done, total=total, ok=ok))
        self._status.setStyleSheet("color: #94A3B8; font-size: 12px;")

    def _on_finished(self, total: int, ok: int) -> None:
        self.btn_run.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self._btn_export.setEnabled(True)
        stopped = self._stop_flag
        key = "tools_batch_stopped" if stopped else "tools_batch_done"
        msg = self._tr(key, ok=ok, total=total)
        color = "#22C55E" if ok == total and not stopped else "#F59E0B"
        self._status.setText(msg)
        self._status.setStyleSheet(f"color: {color}; font-size: 12px;")

    # ── Export ────────────────────────────────────────────────────

    def _on_export(self) -> None:
        path, chosen = QFileDialog.getSaveFileName(
            self, self._tr("tools_batch_dlg_save"),
            "ip_check_results",
            self._tr("tools_batch_dlg_save_filter"),
        )
        if not path:
            return
        try:
            if chosen.startswith("Excel") or path.lower().endswith(".xlsx"):
                if not path.lower().endswith(".xlsx"):
                    path += ".xlsx"
                self._export_xlsx(path)
            else:
                if not path.lower().endswith(".csv"):
                    path += ".csv"
                self._export_csv(path)
            self._status.setText(self._tr("tools_batch_saved", name=Path(path).name))
            self._status.setStyleSheet("color: #22C55E; font-size: 12px;")
        except Exception as exc:
            self._status.setText(self._tr("tools_batch_err_save", err=exc))
            self._status.setStyleSheet("color: #EF4444; font-size: 12px;")

    def _collect_results(self) -> tuple[list[str], list[list[str]]]:
        """Read current table contents into (headers, data_rows)."""
        cols = self._table.columnCount()
        rows = self._table.rowCount()
        headers = [self._table.horizontalHeaderItem(c).text() for c in range(cols)]
        data = [
            [
                (self._table.item(r, c).text() if self._table.item(r, c) else "")
                for c in range(cols)
            ]
            for r in range(rows)
        ]
        return headers, data

    def _export_csv(self, path: str) -> None:
        headers, data = self._collect_results()
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data)

    def _export_xlsx(self, path: str) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError as exc:
            raise ImportError("Установите openpyxl: pip install openpyxl") from exc
        headers, data = self._collect_results()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IP Check"
        # Header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")
        for c_idx, col in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        # Data rows
        status_col = len(headers) - 2  # "Статус" column (1-based)
        for r_idx, row in enumerate(data, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if c_idx == status_col:
                    if val == "OK":
                        cell.font = Font(color="22C55E", bold=True)
                    elif val == "Timeout":
                        cell.font = Font(color="EF4444")
        # Column widths
        for c_idx, col in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = max(
                len(col) + 2,
                max((len(str(row[c_idx - 1])) for row in data), default=0) + 2,
                10,
            )
        wb.save(path)

    def retranslate(self) -> None:
        self._lbl_step1.setText(f"  1. {self._tr('tools_batch_step1')}")
        self._lbl_step2.setText(f"  2. {self._tr('tools_batch_step2')}")
        self._lbl_step3.setText(f"  3. {self._tr('tools_batch_step3')}")
        self._lbl_step4.setText(f"  4. {self._tr('tools_batch_step4')}")
        self._btn_open.setText(self._tr("tools_batch_open_btn"))
        self._lbl_col.setText(self._tr("tools_batch_col_label"))
        self._lbl_timeout.setText(self._tr("tools_batch_timeout_label"))
        self._lbl_timeout_hint.setText(self._tr("tools_batch_timeout_hint"))
        self._lbl_workers.setText(self._tr("tools_batch_workers_label"))
        self.btn_run.setText(self._tr("tools_batch_run_btn"))
        self.btn_stop.setText(self._tr("tools_batch_stop_btn"))
        self._btn_export.setText(self._tr("tools_batch_save_btn"))
        if not self._rows:
            self._lbl_file.setText(self._tr("tools_batch_no_file"))

    def refresh_theme(self, dark: bool) -> None:
        self._dark = dark
        self._table.setStyleSheet(_TREE_SS_LIGHT if not dark else _TREE_SS_DARK)


# fmt: (group_name,) for headers, (name, icon) for tools
_TOOLS = [
    ("Диагностика",),
    ("Ping",           "\u25ce"),   # ◎  target
    ("Traceroute",     "\u21ac"),   # ↬  path
    ("DNS Lookup",     "\u2315"),   # ⌕  search
    ("HTTP Check",     "\u21d7"),   # ⇗  request
    ("SSL Cert",       "\u26BF"),   # ⚿  lock
    ("Локальная сеть",),
    ("tools_nav_adapters", "\u2637"),   # ☷  layers
    ("Netstat",        "\u21c4"),   # ⇄  exchange
    ("ARP",            "\u2237"),   # ∷  table
    ("Routes",         "\u21e2"),   # ⇢  route
    ("Signal Monitor", "\u25d4"),   # ◔  signal
    ("Утилиты",),
    ("Port Scan",      "\u22a1"),   # ⊡  scan
    ("DNS Cache",      "\u2338"),   # ⌸  cache
    ("Subnet Calc",    "\u229e"),   # ⊞  grid
    ("tools_nav_ip_batch", "\u2317"),   # ⌗  grid/list
]


class _ToolNavItem(QFrame):
    """Sidebar nav row: accent-bar | icon | name."""

    clicked_sig = Signal()

    def __init__(self, icon: str, name: str) -> None:
        super().__init__()
        self.setObjectName("ToolNavItem")
        self.setFixedHeight(32)
        self.setCursor(Qt.CursorShape.PointingHandCursor)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Left accent bar
        self._bar = QFrame()
        self._bar.setObjectName("ToolNavBar")
        self._bar.setFixedWidth(3)
        lay.addWidget(self._bar)

        # Icon label — fixed width so text always aligns
        ico_lbl = QLabel(icon)
        ico_lbl.setObjectName("ToolNavIcon")
        ico_lbl.setFixedWidth(26)
        ico_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(ico_lbl)

        self._name = QLabel(name)
        self._name.setObjectName("ToolNavText")
        lay.addWidget(self._name, 1)

    def set_name(self, name: str) -> None:
        self._name.setText(name)

    def set_active(self, active: bool) -> None:
        state = "true" if active else "false"
        for w in (self, self._bar, self._name):
            w.setProperty("active", state)
            w.style().unpolish(w)
            w.style().polish(w)

    def mousePressEvent(self, event) -> None:
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked_sig.emit()
        super().mousePressEvent(event)


class ToolsPage(QWidget):

    def __init__(self, container: "ServiceContainer") -> None:
        super().__init__()
        self._container = container
        self._dark = True
        self._build()

    def _build(self) -> None:
        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        sidebar = QFrame()
        sidebar.setObjectName("ToolsSidebar")
        sidebar.setFixedWidth(190)
        sb_lay = QVBoxLayout(sidebar)
        sb_lay.setContentsMargins(0, 16, 6, 14)
        sb_lay.setSpacing(1)

        self._nav_items: list[_ToolNavItem] = []
        self._item_panel_map: list[int] = []
        self._group_labels: list[QLabel] = []  # refs for retranslate
        self._nav_i18n: list[tuple[_ToolNavItem, str]] = []  # (item, key) for translatable names
        _i18n_build = self._container.i18n
        panel_idx = 0
        first = True
        for entry in _TOOLS:
            if len(entry) == 1:
                if panel_idx > 0:
                    div = QFrame()
                    div.setObjectName("ToolDivider")
                    div.setFixedHeight(1)
                    sb_lay.addSpacing(6)
                    sb_lay.addWidget(div)
                    sb_lay.addSpacing(4)
                grp = QLabel(entry[0])
                grp.setObjectName("ToolGroupLabel")
                self._group_labels.append(grp)
                sb_lay.addWidget(grp)
                sb_lay.addSpacing(3)
            else:
                name, icon = entry
                # If name is an i18n key, translate it
                display_name = _i18n_build.get(name) if name.startswith("tools_") else name
                item = _ToolNavItem(icon, display_name)
                if name.startswith("tools_"):
                    self._nav_i18n.append((item, name))
                item.set_active(first)
                item.clicked_sig.connect(lambda p=panel_idx: self._switch_tool(p))
                sb_lay.addWidget(item)
                self._nav_items.append(item)
                self._item_panel_map.append(panel_idx)
                panel_idx += 1
                first = False

        sb_lay.addStretch(1)
        root.addWidget(sidebar)

        self._stack = QStackedWidget()
        self._stack.setObjectName("ToolsStack")

        # Order must match tool entries in _TOOLS (excluding group headers)
        _i18n = self._container.i18n
        _runner = self._container.process_runner
        self._panels: list = [
            _PingPanel(self._dark, runner=_runner),                              # Диагностика
            _TraceroutePanel(self._dark, i18n=_i18n, runner=_runner),
            _DnsPanel(self._dark, i18n=_i18n, runner=_runner),
            _HttpCheckPanel(self._dark),
            _SslPanel(self._dark, i18n=_i18n),
            _IpconfigPanel(self._dark, i18n=_i18n, runner=_runner),              # Локальная сеть
            _NetstatPanel(self._dark, i18n=_i18n, runner=_runner),
            _ArpPanel(self._dark, i18n=_i18n, runner=_runner),
            _RouteTablePanel(self._dark, i18n=_i18n, runner=_runner),
            _SignalMonitorPanel(self._dark, i18n=_i18n, runner=_runner),
            _PortScanPanel(self._dark, i18n=_i18n),                              # Утилиты
            _DnsCachePanel(self._dark, i18n=_i18n, runner=_runner),
            _SubnetCalcPanel(self._dark, i18n=_i18n),
            _IpBatchPanel(self._dark, i18n=_i18n, runner=_runner),
        ]
        for panel in self._panels:
            self._stack.addWidget(panel)

        root.addWidget(self._stack, 1)
        self._switch_tool(0)

    def _switch_tool(self, idx: int) -> None:
        self._stack.setCurrentIndex(idx)
        for item, p_idx in zip(self._nav_items, self._item_panel_map):
            item.set_active(p_idx == idx)

    def refresh_theme(self, dark_mode: bool) -> None:
        self._dark = dark_mode
        for panel in self._panels:
            panel.refresh_theme(dark_mode)

    def retranslate(self) -> None:
        i18n = self._container.i18n
        # Group labels sidebar
        group_keys = ["tools_group_diagnostics", "tools_group_local", "tools_group_utils"]
        for lbl, key in zip(self._group_labels, group_keys):
            lbl.setText(i18n.get(key))
        # Translatable nav item names
        for item, key in self._nav_i18n:
            item.set_name(i18n.get(key))
        # Buttons in all panels
        run_text   = i18n.get("tools_btn_run")
        stop_text  = i18n.get("tools_btn_stop")
        clear_text = i18n.get("tools_btn_clear")
        for panel in self._panels:
            if hasattr(panel, "btn_run"):
                panel.btn_run.setText(run_text)
            if hasattr(panel, "btn_stop"):
                panel.btn_stop.setText(stop_text)
            if hasattr(panel, "btn_clear"):
                panel.btn_clear.setText(clear_text)
            if hasattr(panel, "retranslate"):
                panel.retranslate()

